#!/usr/bin/env python3
"""Independent schedule validator for RC9.2 artifacts.

This script does not trust optimizer audit metrics.  It reads the input contract,
re-reads the exported schedule/break tables, and independently recomputes key
hard rules, interval coverage, language minima, break structure, rest, OFF shape,
and cyclic next-Sunday coverage.
"""
from __future__ import annotations
import argparse, csv, importlib.util, json, math, re, sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook

DAYS = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

def norm(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip()).casefold()

def hhmm_to_min(v: Any) -> Optional[int]:
    if v is None or str(v).strip() == "": return None
    if hasattr(v, "hour") and hasattr(v, "minute"): return int(v.hour) * 60 + int(v.minute)
    text=str(v).strip()
    m=re.search(r"(?<!\d)(\d{1,2}):(\d{2})(?:\s*([AP]M))?", text, flags=re.I)
    if not m: return None
    h=int(m.group(1)); minute=int(m.group(2)); ap=(m.group(3) or "").upper()
    if ap:
        h%=12
        if ap=="PM": h+=12
    if h==24 and minute==0: h=0
    if not (0<=h<24 and 0<=minute<60): return None
    return h*60+minute

def load_engine(engine_path: Path):
    sys.path.insert(0, str(engine_path.parent))
    spec=importlib.util.spec_from_file_location("validator_contract_parser", engine_path)
    mod=importlib.util.module_from_spec(spec); sys.modules[spec.name]=mod; spec.loader.exec_module(mod)
    return mod

def find_header(ws, required=("sf name",)) -> int:
    for r in range(1, min(ws.max_row, 20)+1):
        vals={norm(ws.cell(r,c).value) for c in range(1,ws.max_column+1)}
        if any(req in vals for req in required): return r
    raise ValueError(f"Could not find header in {ws.title}")

def parse_output_schedule(path: Path, expected_names: List[str]) -> Tuple[Dict[str,List[str]], Dict[str,str]]:
    wb=load_workbook(path,data_only=False,read_only=False)
    ws=wb["Schedule"] if "Schedule" in wb.sheetnames else wb["Final Schedule"]
    header=find_header(ws)
    headers={norm(ws.cell(header,c).value):c for c in range(1,ws.max_column+1)}
    name_col=next((c for h,c in headers.items() if h in {"sf name","associate","associate name","name"}),None)
    lang_col=next((c for h,c in headers.items() if h=="language"),None)
    # Day columns are not always labelled on the same row as the name column.
    # Exported schedules carry the day NAMES on the banner row and the calendar
    # DATES on the row holding "SF Name", so locking to the name row found no
    # day columns and the validator raised, aborting before it evaluated a
    # single rule.  Search the name row first, then its neighbours, accepting
    # short or full day names - the same convention the engine's own
    # _day_columns uses.
    day_cols=None
    for probe in (header, header-1, header+1, header-2):
        if probe<1 or probe>ws.max_row: continue
        probe_headers={norm(ws.cell(probe,c).value):c for c in range(1,ws.max_column+1)}
        candidate={}
        for d,full in zip(DAYS,("sunday","monday","tuesday","wednesday","thursday","friday","saturday")):
            match=next((c for h,c in probe_headers.items()
                        if h==norm(d) or h.startswith(norm(d)) or h.startswith(full)),None)
            candidate[d]=match
        if all(c is not None for c in candidate.values()):
            day_cols=candidate; break
    if name_col is None or day_cols is None:
        raise ValueError(
            "Output schedule is missing name/day columns "
            f"(name_col={name_col}, searched rows {max(1,header-2)}-{header+1} of sheet {ws.title!r})")
    assignments={}; languages={}
    for r in range(header+1,ws.max_row+1):
        name=str(ws.cell(r,name_col).value or "").strip()
        if not name: continue
        assignments[norm(name)]=[str(ws.cell(r,day_cols[d]).value or "").strip() for d in DAYS]
        languages[norm(name)]=str(ws.cell(r,lang_col).value or "").strip() if lang_col else ""
    wb.close()
    return assignments,languages

def parse_breaks(path: Path) -> Tuple[List[Dict[str,Any]], bool]:
    wb=load_workbook(path,data_only=True,read_only=True)
    sheet=next((wb[s] for s in ["Break Schedule Active","Break Schedule"] if s in wb.sheetnames),None)
    if sheet is None:
        wb.close(); return [],False
    rows=list(sheet.iter_rows(values_only=True))
    if not rows:
        wb.close(); return [],False
    headers={norm(v):i for i,v in enumerate(rows[0])}
    out=[]; before_only=False
    for rr in rows[1:]:
        status=str(rr[headers.get("status",-1)] or "") if headers.get("status") is not None else ""
        if "not assigned in before-break" in status.casefold(): before_only=True
        name=str(rr[headers.get("associate",0)] or "").strip() if rr else ""
        day=str(rr[headers.get("day",1)] or "").strip() if len(rr)>1 else ""
        if not name or norm(day) not in {norm(d) for d in DAYS}: continue
        out.append({
            "associate":name,"day":day,"shift":str(rr[headers.get("shift",2)] or ""),
            "type":str(rr[headers.get("break type",3)] or ""),"start":rr[headers.get("start",4)],
            "duration":int(float(rr[headers.get("duration minutes",5)] or 0)),"status":status,
        })
    wb.close(); return out,before_only

def max_gap_run(flags: List[Optional[bool]], intervals_per_day: int) -> int:
    best=0
    for d in range(7):
        run=0
        for flag in flags[d*intervals_per_day:(d+1)*intervals_per_day]:
            if flag is None: run=0
            elif flag: run+=1; best=max(best,run)
            else: run=0
    return best

def validate(input_path: Path, output_path: Path, engine_path: Path) -> Dict[str,Any]:
    eng=load_engine(engine_path)
    parsed=eng.parse_input(input_path)
    assignments,_=parse_output_schedule(output_path,[a.name for a in parsed.associates])
    breaks,before_only=parse_breaks(output_path)
    shift_map={norm(s.label):s for s in parsed.shifts}
    failures=[]; warnings=[]
    matrix=[]
    for a,assoc in enumerate(parsed.associates):
        row=assignments.get(norm(assoc.name))
        if row is None:
            failures.append({"type":"MISSING_ASSOCIATE","associate":assoc.name}); row=[""]*7
        matrix.append(row)
    extra=sorted(set(assignments)-{norm(a.name) for a in parsed.associates})
    if extra: warnings.append({"type":"EXTRA_OUTPUT_ASSOCIATES","count":len(extra),"examples":extra[:10]})

    # Assignment legality, fixed/leave/OFF, OFF shape, shift variety and rest.
    for a,assoc in enumerate(parsed.associates):
        row=matrix[a]; off_days=[]; shift_days=[]
        for d,value in enumerate(row):
            kind="off" if norm(value)=="off" else "leave" if norm(value) in {"leave","pto","vacation"} else "shift" if norm(value) in shift_map else "blank"
            if kind=="blank": failures.append({"type":"UNKNOWN_OR_BLANK_ASSIGNMENT","associate":assoc.name,"day":DAYS[d],"value":value})
            if kind=="off": off_days.append(d)
            if kind=="shift": shift_days.append((d,shift_map[norm(value)]))
            pref_kind=eng.preference_kind(assoc.preferences[d] if d<len(assoc.preferences) else "")
            fixed_kind=eng.preference_kind(assoc.fixed_schedule[d] if d<len(assoc.fixed_schedule) else "")
            if parsed.leave_enabled and (pref_kind=="leave" or fixed_kind=="leave") and kind!="leave":
                failures.append({"type":"LEAVE_VIOLATION","associate":assoc.name,"day":DAYS[d],"actual":value})
            if parsed.hard_off and (pref_kind=="off" or fixed_kind=="off") and kind!="off":
                failures.append({"type":"HARD_OFF_VIOLATION","associate":assoc.name,"day":DAYS[d],"actual":value})
            if parsed.fixed_enabled and fixed_kind=="shift" and norm(value)!=norm(assoc.fixed_schedule[d]):
                failures.append({"type":"FIXED_SHIFT_VIOLATION","associate":assoc.name,"day":DAYS[d],"expected":assoc.fixed_schedule[d],"actual":value})
        # Contract constant shared with the engine, not an independent judgement.
        # Independence means recomputing coverage from the exported workbook - it
        # does not mean re-guessing what the contract says a long shift is.  This
        # was a local 660 while the engine used 630, so a shift in [630, 660) was
        # long to the engine and short to the validator, producing a false
        # OFF_COUNT_VIOLATION that would block an otherwise valid release.
        long_mode=any(s.duration_min>=eng.LONG_SHIFT_MIN_DURATION_MIN for _,s in shift_days)
        expected_off=3 if long_mode else 2
        if parsed.strict_off and len(off_days)!=expected_off:
            failures.append({"type":"OFF_COUNT_VIOLATION","associate":assoc.name,"expected":expected_off,"actual":len(off_days),"off_days":[DAYS[d] for d in off_days]})
        if parsed.strict_off and not parsed.separate_off_days:
            consecutive=any(d in off_days and (d+1)%7 in off_days for d in range(7))
            if not consecutive: failures.append({"type":"CONSECUTIVE_OFF_VIOLATION","associate":assoc.name,"off_days":[DAYS[d] for d in off_days]})
        distinct=len({norm(s.label) for _,s in shift_days})
        if distinct>parsed.max_different_shifts:
            failures.append({"type":"MAX_SHIFT_VARIETY_VIOLATION","associate":assoc.name,"actual":distinct,"maximum":parsed.max_different_shifts})
        for d in range(6):
            left=shift_map.get(norm(row[d])); right=shift_map.get(norm(row[d+1]))
            if left and right and not eng.rest_compatible(left,right,parsed.rest_gap_hours):
                failures.append({"type":"REST_VIOLATION","associate":assoc.name,"from_day":DAYS[d],"to_day":DAYS[d+1],"from":left.label,"to":right.label})
        sat=shift_map.get(norm(row[6])); sun=shift_map.get(norm(row[0]))
        if sat and sun and not eng.rest_compatible(sat,sun,parsed.rest_gap_hours):
            failures.append({"type":"CYCLIC_REST_VIOLATION","associate":assoc.name,"from":"Sat","to":"Sun","from_shift":sat.label,"to_shift":sun.label})
        if sun and not eng.previous_saturday_compatible(assoc.previous_saturday,sun,parsed.rest_gap_hours):
            failures.append({"type":"PREVIOUS_SATURDAY_REST_VIOLATION","associate":assoc.name,"previous_saturday":assoc.previous_saturday,"sunday":sun.label})

    # Build shift occurrences at quarter level.
    horizon=7*96
    before=[[] for _ in range(horizon)]
    for a,assoc in enumerate(parsed.associates):
        for d,value in enumerate(matrix[a]):
            shift=shift_map.get(norm(value))
            if not shift: continue
            start=d*96+shift.start_min//15
            for q in range(start,start+shift.duration_q):
                if 0<=q<horizon: before[q].append(a)
        # previous Saturday spill into current Sunday
        pshift=shift_map.get(norm(assoc.previous_saturday))
        if pshift:
            start=-96+pshift.start_min//15
            for q in range(start,start+pshift.duration_q):
                if 0<=q<96: before[q].append(a)

    break_qslots=defaultdict(set); break_fail=[]
    by_cell=defaultdict(list)
    for br in breaks:
        aidx=next((i for i,a in enumerate(parsed.associates) if norm(a.name)==norm(br['associate'])),None)
        didx=next((i for i,d in enumerate(DAYS) if norm(d)==norm(br['day'])),None)
        if aidx is None or didx is None: break_fail.append({"type":"UNKNOWN_BREAK_ASSOCIATE_OR_DAY","row":br}); continue
        shift=shift_map.get(norm(matrix[aidx][didx])); start_clock=hhmm_to_min(br['start'])
        if shift is None or start_clock is None or br['duration']<=0:
            break_fail.append({"type":"INVALID_BREAK_ROW","row":br}); continue
        rel=start_clock-shift.start_min
        if rel<0: rel+=1440
        if rel<0 or rel+br['duration']>shift.duration_min:
            break_fail.append({"type":"BREAK_OUTSIDE_SHIFT","associate":parsed.associates[aidx].name,"day":DAYS[didx],"break":br})
        abs_start=didx*96+(shift.start_min+rel)//15
        duration_q=math.ceil(br['duration']/15)
        for q in range(abs_start,abs_start+duration_q):
            if 0<=q<horizon:
                if q in break_qslots[aidx]: break_fail.append({"type":"OVERLAPPING_BREAK","associate":parsed.associates[aidx].name,"day":DAYS[didx],"qslot":q})
                break_qslots[aidx].add(q)
        by_cell[aidx,didx].append(br)
    failures.extend(break_fail)
    if not before_only:
        expected_segments=[(q*15,label) for q,label in parsed.break_segments_q]
        for a,assoc in enumerate(parsed.associates):
            for d,value in enumerate(matrix[a]):
                if norm(value) not in shift_map: continue
                cell=by_cell.get((a,d),[])
                durations=sorted(int(r['duration']) for r in cell)
                expected=sorted(m for m,_ in expected_segments)
                if durations!=expected:
                    failures.append({"type":"BREAK_SEGMENT_COUNT_OR_DURATION","associate":assoc.name,"day":DAYS[d],"expected":expected,"actual":durations})

    after=[]
    for q,covers in enumerate(before):
        after.append([a for a in covers if q not in break_qslots[a]])

    # Coverage and language.
    interval_rows=[]; zero=[]; language_gaps=[]; opening_gaps=[]; floor_flags=[]
    before100=before90=before80=after100=after90=after80=before_target=after_target=before_floor=after_floor=0
    for d in range(7):
        openings=set(eng.opening_intervals_for_day(parsed,d))
        for i in range(parsed.intervals_per_day):
            if not parsed.active[d][i]:
                floor_flags.append(None); continue
            req=float(parsed.requirements[d][i] or 0.0); eff=1-float(parsed.shrinkage[d][i]); qpi=parsed.qslots_per_interval
            bvals=[]; avals=[]
            for q in range(qpi):
                slot=d*96+i*qpi+q
                b=len(before[slot]); a=len(after[slot]); bvals.append(b); avals.append(a)
                minute=i*parsed.interval_minutes+q*15
                if a<=0: zero.append({"day":DAYS[d],"time":eng.hhmm(minute)})
                if parsed.opening_guard_enabled and i in openings and a<parsed.opening_minimum:
                    opening_gaps.append({"day":DAYS[d],"time":eng.hhmm(minute),"actual":a,"minimum":parsed.opening_minimum})
                for rule in parsed.language_rules:
                    if not rule.active or not rule.contains_minute(minute): continue
                    eligible=sum(1 for ai in after[slot] if norm(parsed.associates[ai].language) in rule.eligible_languages)
                    if eligible<rule.minimum:
                        language_gaps.append({"day":DAYS[d],"time":eng.hhmm(minute),"group":rule.group,"minimum":rule.minimum,"actual":eligible})
            be=sum(bvals)*eff/qpi; ae=sum(avals)*eff/qpi
            bp=be/req if req>0 else 1.0; ap=ae/req if req>0 else 1.0
            before100+=bp>=1-1e-9; before90+=bp>=.9-1e-9; before80+=bp>=.8-1e-9
            after100+=ap>=1-1e-9; after90+=ap>=.9-1e-9; after80+=ap>=.8-1e-9
            before_target+=bp+1e-9>=parsed.target_ratio; after_target+=ap+1e-9>=parsed.target_ratio
            before_floor+=bp+1e-9>=parsed.floor_ratio; after_floor+=ap+1e-9>=parsed.floor_ratio
            floor_flags.append(ap+1e-9<parsed.floor_ratio)
            # Same ceil tolerance the engine applies.  Without it a requirement
            # that is integral in exact arithmetic but marginally above integral
            # in floating point rounds up by a whole associate, overstating
            # unavoidable staffing and understating avoidable overage - a silent
            # disagreement with the engine on the exact metric under review.
            unavoidable=math.ceil(req*parsed.target_ratio/max(eff,1e-9)-eng.OVERAGE_CEIL_TOLERANCE)*eff
            avoid=max(0.0,ae-unavoidable)
            interval_rows.append({"day":DAYS[d],"interval":eng.hhmm(i*parsed.interval_minutes),"required":req,"before_effective":be,"after_effective":ae,"before_pct":bp,"after_pct":ap,"avoidable_overage_fte":avoid})
    active=len(interval_rows); severe_threshold=max(0,parsed.floor_ratio-.10)
    severe=sum(1 for r in interval_rows if r['after_pct']+1e-9<severe_threshold)
    maxrun=max_gap_run(floor_flags,parsed.intervals_per_day)
    overages=[r['avoidable_overage_fte'] for r in interval_rows]
    mean=sum(overages)/max(1,len(overages)); variance=sum((x-mean)**2 for x in overages)/max(1,len(overages))

    # Next Sunday independent boundary check.
    next_rows=[]
    for i in range(parsed.intervals_per_day):
        if not parsed.active[0][i]: continue
        req=float(parsed.requirements[0][i] or 0); eff=1-float(parsed.shrinkage[0][i]); vals=[]
        for q in range(parsed.qslots_per_interval):
            pseudo=7*96+i*parsed.qslots_per_interval+q; count=0
            for a,assoc in enumerate(parsed.associates):
                sat=shift_map.get(norm(matrix[a][6])); sun=shift_map.get(norm(matrix[a][0]))
                if sat:
                    st=6*96+sat.start_min//15
                    if st<=pseudo<st+sat.duration_q: count+=1
                if sun:
                    st=7*96+sun.start_min//15
                    if st<=pseudo<st+sun.duration_q: count+=1
            vals.append(count)
        ae=sum(vals)*eff/max(1,len(vals)); pct=ae/req if req>0 else 1
        next_rows.append({"interval":eng.hhmm(i*parsed.interval_minutes),"after_effective":ae,"pct":pct})

    failures.extend({"type":"ZERO_STAFF_ACTIVE","detail":row} for row in zero)
    failures.extend({"type":"LANGUAGE_MINIMUM","detail":row} for row in language_gaps)
    failures.extend({"type":"OPENING_MINIMUM","detail":row} for row in opening_gaps)
    metrics={
        "active_intervals":active,"before100":int(before100),"before90":int(before90),"before80":int(before80),
        "after100":int(after100),"after90":int(after90),"after80":int(after80),
        "before_target":int(before_target),"after_target":int(after_target),"before_floor":int(before_floor),"after_floor":int(after_floor),
        "floor_gaps":active-int(after_floor),"severe_floor_gaps":severe,"max_consecutive_floor_gaps":maxrun,
        "zero_staffed_active_quarters":len(zero),"language_gap_count":len(language_gaps),"opening_gap_count":len(opening_gaps),
        "avoidable_overage_fte_sum":sum(overages),"avoidable_overage_peak_fte":max(overages,default=0),
        "avoidable_overage_variance_fte":variance,"avoidable_overage_stddev_fte":math.sqrt(max(0,variance)),
        "next_sunday_target_hits":sum(r['pct']+1e-9>=parsed.target_ratio for r in next_rows),
        "next_sunday_floor_hits":sum(r['pct']+1e-9>=parsed.floor_ratio for r in next_rows),
    }
    return {
        "schema_version":1,"validator":"RC9.2_INDEPENDENT_VALIDATOR","input":str(input_path),"output":str(output_path),
        "artifact_role":"BEST_BEFORE_BREAKS" if before_only else "FINAL_AFTER_BREAKS",
        "status":"PASS" if not failures else "FAIL","hard_fail_count":len(failures),"warning_count":len(warnings),
        "metrics":metrics,"failures":failures[:500],"warnings":warnings[:500],"interval_rows":interval_rows,"next_sunday_rows":next_rows,
        "evidence_note":"Metrics were recomputed from workbook schedule/break cells and input contract; optimizer audit values were not used."
    }

def main():
    p=argparse.ArgumentParser()
    p.add_argument('--input',type=Path,required=True)
    p.add_argument('--output',type=Path,required=True)
    p.add_argument('--engine',type=Path,default=Path(__file__).resolve().parents[1]/'_tools/l632_universal_scheduler.py')
    p.add_argument('--json-out',type=Path)
    p.add_argument('--csv-out',type=Path)
    args=p.parse_args()
    result=validate(args.input,args.output,args.engine)
    jout=args.json_out or args.output.with_name(args.output.stem+'_INDEPENDENT_VALIDATION.json')
    jout.parent.mkdir(parents=True,exist_ok=True); jout.write_text(json.dumps(result,indent=2,default=str),encoding='utf-8')
    cout=args.csv_out or args.output.with_name(args.output.stem+'_INDEPENDENT_VALIDATION.csv')
    with cout.open('w',newline='',encoding='utf-8') as f:
        w=csv.writer(f); w.writerow(['Metric','Value']);
        for k,v in result['metrics'].items(): w.writerow([k,v])
        w.writerow(['status',result['status']]); w.writerow(['hard_fail_count',result['hard_fail_count']]); w.writerow(['warning_count',result['warning_count']])
    print(json.dumps({k:result[k] for k in ['status','hard_fail_count','warning_count','artifact_role','metrics']},indent=2))
    return 0 if result['status']=='PASS' else 2
if __name__=='__main__': raise SystemExit(main())
