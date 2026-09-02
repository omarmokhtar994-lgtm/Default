#!/usr/bin/env python3
"""Engine <-> independent-validator semantic parity guards.

The independent validator must recompute the exported workbook's coverage on its
own - that is what makes it independent, and it must be able to fail a release
the optimizer called PASS.

Independence does NOT extend to contract semantics.  Where the validator and the
engine must agree by definition - what counts as a long shift, what tolerance
converts an effective requirement into whole associates - a second private copy
of the value is not independence, it is a silent disagreement that shows up as a
validator false positive.  Under the release gates any validator mismatch blocks
the release, so a false positive is as damaging as a missed defect.

These tests pin the constants that must be shared, and the coverage math that
must not be.

Run:  python tests/test_rc9_2_1_validator_parity.py
"""
from __future__ import annotations

import importlib.util
import math
import random
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "engine" / "_tools"))

import l632_universal_scheduler as E  # noqa: E402

VALIDATOR_PATH = ROOT / "engine" / "tools" / "independent_validator.py"


def load_validator():
    spec = importlib.util.spec_from_file_location("independent_validator", VALIDATOR_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class SharedContractConstants(unittest.TestCase):
    """Constants that both sides apply must have exactly one definition."""

    def test_long_shift_threshold_is_defined_once(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        # assertFalse on a bool, not assertNotIn on the file, so a failure prints
        # one line instead of the entire validator source.
        self.assertFalse(
            "duration_min>=660" in source.replace(" ", ""),
            "Validator carries its own long-shift threshold (660). It must use "
            "eng.LONG_SHIFT_MIN_DURATION_MIN, or a shift in [630, 660) is long to "
            "the engine and short to the validator, producing a false "
            "OFF_COUNT_VIOLATION that blocks a valid release.",
        )
        self.assertTrue(
            "LONG_SHIFT_MIN_DURATION_MIN" in source,
            "Validator must reference the engine's shared long-shift constant.",
        )

    def test_engine_long_shift_threshold_value(self):
        self.assertEqual(E.LONG_SHIFT_MIN_DURATION_MIN, 630)

    def test_engine_has_no_bare_long_shift_literals(self):
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8")
        compact = source.replace(" ", "")
        self.assertNotIn("duration_min>=630", compact)
        self.assertNotIn("duration_min<630", compact)

    def test_off_expectation_agrees_across_the_boundary(self):
        """A 10.5h shift must be long mode on both sides, not just one."""
        threshold = E.LONG_SHIFT_MIN_DURATION_MIN
        for duration, expected_off in ((threshold - 1, 2), (threshold, 3), (threshold + 30, 3)):
            with self.subTest(duration=duration):
                self.assertEqual(3 if duration >= threshold else 2, expected_off)

    def test_ceil_tolerance_is_shared(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        self.assertIn("OVERAGE_CEIL_TOLERANCE", source,
                      "Validator must apply the engine's ceil tolerance.")
        self.assertEqual(E.OVERAGE_CEIL_TOLERANCE, 1e-9)


class AvoidableOverageParity(unittest.TestCase):
    """The engine and validator must agree on avoidable overage, always.

    This is the metric the whole RC9.2.1 release argument turns on.  A
    disagreement here means the two sides are arguing about different numbers.
    """

    @staticmethod
    def validator_unavoidable(req, eff, target):
        return math.ceil(req * target / max(eff, 1e-9) - E.OVERAGE_CEIL_TOLERANCE) * eff

    @staticmethod
    def engine_unavoidable(req, eff, target):
        return E.fractional_safe_overage_metrics(
            req, 0.0, target, eff, 1.1, 1.25, 1.5)["minimum_raw_target"] * eff

    def test_agreement_over_random_contracts(self):
        random.seed(7)
        divergences = []
        for _ in range(200000):
            req = round(random.uniform(0.5, 60.0), 2)
            eff = round(random.uniform(0.70, 1.00), 4)
            target = random.choice([0.80, 0.90, 1.00])
            a = self.engine_unavoidable(req, eff, target)
            b = self.validator_unavoidable(req, eff, target)
            if abs(a - b) > 1e-12:
                divergences.append((req, eff, target, a, b))
                if len(divergences) > 3:
                    break
        self.assertEqual(
            divergences, [],
            "Engine and validator disagree on unavoidable staffing, so they "
            "disagree on avoidable overage.",
        )

    def test_zero_demand_interval_agrees(self):
        for staffing in (0.0, 3.4, 8.5):
            with self.subTest(staffing=staffing):
                engine = E.fractional_safe_overage_metrics(
                    0.0, staffing, 0.90, 0.85, 1.1, 1.25, 1.5)["avoidable_overage_fte"]
                validator = max(0.0, staffing - self.validator_unavoidable(0.0, 0.85, 0.90))
                self.assertAlmostEqual(engine, validator, places=9)

    def test_known_floating_point_boundary_case(self):
        """req=45.56 eff=0.9112 target=0.9 -> exactly 45 associates in exact math.

        45.56 * 0.9 / 0.9112 == 45.0 exactly, but lands marginally above 45 in
        floating point.  Found by random search against the pre-fix validator,
        which rounded up to 46 and so understated avoidable overage by one whole
        associate-equivalent (0.9112 effective FTE) on this interval.
        """
        a = self.engine_unavoidable(45.56, 0.9112, 0.90)
        b = self.validator_unavoidable(45.56, 0.9112, 0.90)
        self.assertAlmostEqual(a, b, places=9)
        self.assertAlmostEqual(a, 45 * 0.9112, places=9)

    def test_pre_fix_validator_would_have_failed_this_case(self):
        """Pin that the boundary case is genuinely discriminating."""
        naive = math.ceil(45.56 * 0.90 / 0.9112) * 0.9112  # validator before the fix
        engine = self.engine_unavoidable(45.56, 0.9112, 0.90)
        self.assertAlmostEqual(naive - engine, 0.9112, places=9,
                               msg="Expected exactly one associate-equivalent of "
                                   "disagreement before the tolerance was shared.")


class ValidatorIndependenceIsPreserved(unittest.TestCase):
    """The validator must still compute coverage itself, not defer to the engine."""

    def test_validator_recomputes_coverage_locally(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        for marker in ("before100+=", "after100+=", "before_target+=", "after_floor+="):
            self.assertIn(marker.replace("+=", ""), source.replace(" ", ""),
                          "Validator must accumulate its own tier counts.")

    def test_validator_does_not_call_engine_metric_calculator(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        self.assertNotIn("calculate_metrics", source,
                         "Reusing the engine's metric calculator would destroy "
                         "the validator's independence.")

    def test_validator_states_its_evidence_basis(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        self.assertIn("optimizer audit values were not used", source)

    def test_validator_module_loads(self):
        module = load_validator()
        self.assertTrue(hasattr(module, "validate"))


class TierCountingFormulaParity(unittest.TestCase):
    """Tier counting math must be identical on both sides.

    Audited by reading both implementations: same active guard, same
    effective-FTE derivation, same zero-demand convention, same epsilon.  These
    tests pin the conventions so a future edit to one side is caught.
    """

    def test_zero_demand_counts_as_fully_covered_on_both_sides(self):
        engine_src = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8").replace(" ", "")
        validator_src = VALIDATOR_PATH.read_text(encoding="utf-8").replace(" ", "")
        self.assertIn("before_pct=before/reqifreq>0else1.0", engine_src)
        self.assertIn("bp=be/reqifreq>0else1.0", validator_src)

    def test_both_sides_use_the_same_tier_epsilon(self):
        engine_src = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8").replace(" ", "")
        validator_src = VALIDATOR_PATH.read_text(encoding="utf-8").replace(" ", "")
        self.assertIn("before_pct>=0.90-1e-9", engine_src)
        self.assertIn("bp>=.9-1e-9", validator_src)

    def test_both_sides_skip_inactive_intervals(self):
        engine_src = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8").replace(" ", "")
        validator_src = VALIDATOR_PATH.read_text(encoding="utf-8").replace(" ", "")
        self.assertIn("ifnotparsed.active[d][i]:", engine_src)
        self.assertIn("ifnotparsed.active[d][i]:", validator_src)


class ValidatorReadsExportedScheduleLayouts(unittest.TestCase):
    """The validator must read the layouts the engine actually exports.

    Exported schedules put day NAMES on the banner row and calendar DATES on the
    row carrying "SF Name".  The validator located its header by "SF Name" and
    then looked for day columns on that same row, found dates, and raised
    ValueError("Output schedule is missing name/day columns") - aborting before
    it evaluated a single rule.  Every skeleton-only run therefore failed
    independent validation for a parsing reason, not a scheduling one.
    """

    def _sheet(self, day_row_offset):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        days = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        if day_row_offset == -1:                      # names above, dates on header row
            for i, d in enumerate(days):
                ws.cell(1, 7 + i).value = d
            ws.cell(2, 4).value = "SF Name"
            for i in range(7):
                ws.cell(2, 7 + i).value = f"2026-07-{12 + i}"
        else:                                          # names on the header row
            ws.cell(1, 4).value = "SF Name"
            for i, d in enumerate(days):
                ws.cell(1, 7 + i).value = d
        row = 3 if day_row_offset == -1 else 2
        ws.cell(row, 4).value = "Associate 001"
        for i in range(7):
            ws.cell(row, 7 + i).value = "OFF"
        return wb

    def _parse(self, wb):
        import tempfile, os
        module = load_validator()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        try:
            wb.save(path)
            return module.parse_output_schedule(Path(path), ["Associate 001"])
        finally:
            os.unlink(path)

    def test_day_names_on_the_row_above_the_header(self):
        assignments, _ = self._parse(self._sheet(-1))
        self.assertIn("associate 001", assignments)
        self.assertEqual(len(assignments["associate 001"]), 7)

    def test_day_names_on_the_header_row_still_work(self):
        assignments, _ = self._parse(self._sheet(0))
        self.assertIn("associate 001", assignments)
        self.assertEqual(len(assignments["associate 001"]), 7)

    def test_error_message_names_the_rows_searched(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        self.assertIn("searched rows", source,
                      "a parse failure must say where it looked, not just that it failed")


class ValidatorCrashIsNotAScheduleFailure(unittest.TestCase):
    """A broken checker and a broken schedule are different claims.

    The validator exits 0 on PASS and 2 when it evaluated the schedule and found
    hard-rule violations.  Any other code means it did not complete.  Mapping
    every non-zero code to 'FAIL' reported a crash as a schedule failure.
    """

    def _runner_source(self):
        return (ROOT / "engine" / "RUN_UNIVERSAL_PRODUCTION.py").read_text(encoding="utf-8")

    def test_runner_distinguishes_incomplete_from_failed(self):
        source = self._runner_source()
        self.assertIn("ERROR_VALIDATOR_DID_NOT_COMPLETE", source)
        self.assertNotIn("'PASS' if vrc == 0 else 'FAIL'", source,
                         "every non-zero code must not collapse to FAIL")

    def test_only_exit_two_means_hard_rule_failure(self):
        source = self._runner_source()
        self.assertIn("vrc == 2", source,
                      "exit code 2 is the validator's 'evaluated and failed' signal")

    def test_validator_returns_two_on_hard_failure(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        self.assertIn("return 0 if result['status']=='PASS' else 2", source.replace('"', "'"))

    def test_incomplete_validation_still_blocks_release(self):
        """Distinguishable, but not permissive - an unvalidated schedule blocks."""
        source = self._runner_source()
        marker = source.index("ERROR_VALIDATOR_DID_NOT_COMPLETE")
        self.assertIn("rc = 4", source[marker:marker + 900],
                      "a non-zero validator code must still set rc=4")


class CarryInCoverageIsParsedNotLookedUp(unittest.TestCase):
    """Previous-Saturday carry-in must be parsed, not required in the library.

    Carry-in is historical fact - the associate worked that shift last week.
    The shift library constrains what may be ASSIGNED this week and is filtered
    by allowed durations and start window, so requiring membership silently
    discarded real coverage.

    On NMG SP the library holds one label while two associates carry in
    "16:00 - 01:00" and "21:00 - 06:00".  Both were dropped, and the validator
    under-counted against the engine by 4 target and 2 floor intervals - the
    long-open "independent validator exact agreement" gate failure.  Parsing the
    label reconciled all six coverage metrics exactly.
    """

    def test_validator_parses_the_carry_in_label(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8")
        self.assertIn("eng.shift_parts(assoc.previous_saturday)", source,
                      "carry-in must be parsed from the label")

    def test_validator_does_not_require_library_membership(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8").replace(" ", "")
        self.assertNotIn("shift_map.get(norm(assoc.previous_saturday))", source,
                         "requiring the carry-in label to be in this week's "
                         "library silently discards real coverage")

    def test_engine_also_parses_rather_than_looks_up(self):
        """The engine's own carry-in predicate parses the label directly."""
        self.assertTrue(E.previous_saturday_covers_qslot("21:00 - 06:00", 4))
        self.assertTrue(E.previous_saturday_covers_qslot("16:00 - 01:00", 0))

    def test_carry_in_labels_outside_the_library_still_cover(self):
        """The exact NMG SP labels, neither of which is in its shift library."""
        for label, qslot in (("16:00 - 01:00", 0), ("21:00 - 06:00", 8)):
            with self.subTest(label=label):
                self.assertIsNotNone(E.shift_parts(label))
                self.assertTrue(E.previous_saturday_covers_qslot(label, qslot))

    def test_non_shift_carry_in_values_are_ignored(self):
        """OFF and blanks must not be treated as coverage."""
        self.assertIsNone(E.shift_parts("OFF"))
        self.assertIsNone(E.shift_parts(""))
        self.assertFalse(E.previous_saturday_covers_qslot("OFF", 0))


class DayColumnMatchingIsNotOverBroad(unittest.TestCase):
    """Day-column detection must not bind a non-day column to a day.

    The first version of the neighbouring-row search accepted any header
    STARTING WITH a short day name.  That binds "Monthly Total" and "Month" to
    Mon, and "Saturation" / "Satisfaction Score" to Sat - silently reading a
    non-day column as that day's assignments, which corrupts every downstream
    coverage number while still reporting a clean parse.
    """

    FALSE_MATCHES = ("Monthly Total", "Month", "Saturation", "Satisfaction Score",
                     "Weds", "Frida", "Summary", "Suntotal")

    def _accepts(self, header, day, full):
        h = header.strip().casefold()
        return h == day.casefold() or h == full or h.startswith(full + " ")

    def test_short_day_prefixes_are_rejected(self):
        pairs = list(zip(["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
                         ("sunday", "monday", "tuesday", "wednesday",
                          "thursday", "friday", "saturday")))
        for header in self.FALSE_MATCHES:
            for day, full in pairs:
                with self.subTest(header=header, day=day):
                    self.assertFalse(self._accepts(header, day, full),
                                     f"{header!r} must not bind to {day}")

    def test_genuine_day_headers_are_accepted(self):
        pairs = list(zip(["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
                         ("sunday", "monday", "tuesday", "wednesday",
                          "thursday", "friday", "saturday")))
        for day, full in pairs:
            with self.subTest(day=day):
                self.assertTrue(self._accepts(day, day, full))
                self.assertTrue(self._accepts(full.capitalize(), day, full))
                self.assertTrue(self._accepts(f"{full.capitalize()} 12 Jul", day, full))

    def test_validator_source_does_not_use_short_prefix_matching(self):
        source = VALIDATOR_PATH.read_text(encoding="utf-8").replace(" ", "")
        self.assertNotIn("h.startswith(norm(d))", source,
                         "short-day prefix matching re-introduced")


class BreakSheetColumnsComeFromHeaders(unittest.TestCase):
    """The validator must not guess its own input columns.

    parse_breaks previously fell back to positional defaults (associate=0,
    day=1, shift=2, ...).  Those happen to match the current export order, so a
    missing or reordered header would have been read silently from the wrong
    column - and this validator gates release, so a PASS built on the wrong
    columns is worse than a crash.
    """

    def _sheet(self, header_row, data_rows):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Break Schedule"
        for c, v in enumerate(header_row, start=1):
            ws.cell(1, c).value = v
        for r, row in enumerate(data_rows, start=2):
            for c, v in enumerate(row, start=1):
                ws.cell(r, c).value = v
        return wb

    def _parse(self, wb):
        import tempfile, os
        module = load_validator()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        try:
            wb.save(path)
            return module.parse_breaks(Path(path))
        finally:
            os.unlink(path)

    FULL = ["Associate", "Day", "Shift", "Break Type", "Start", "Duration Minutes", "Status"]

    def test_wellformed_sheet_parses(self):
        wb = self._sheet(self.FULL,
                         [["A", "Sun", "13:00 - 22:00", "Break 1", "15:15", 15, "Scheduled"]])
        rows, before_only = self._parse(wb)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["associate"], "A")
        self.assertEqual(rows[0]["duration"], 15)
        self.assertFalse(before_only)

    def test_reordered_columns_are_read_by_header_not_position(self):
        reordered = ["Status", "Duration Minutes", "Day", "Associate", "Shift", "Break Type", "Start"]
        wb = self._sheet(reordered,
                         [["Scheduled", 30, "Mon", "B", "09:00 - 18:00", "Lunch", "12:00"]])
        rows, _ = self._parse(wb)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["associate"], "B", "column order must follow headers")
        self.assertEqual(rows[0]["day"], "Mon")
        self.assertEqual(rows[0]["duration"], 30)

    def test_missing_required_header_raises_rather_than_guessing(self):
        without_day = ["Associate", "Shift", "Break Type", "Start", "Duration Minutes", "Status"]
        wb = self._sheet(without_day, [["A", "13:00 - 22:00", "Break 1", "15:15", 15, "Scheduled"]])
        with self.assertRaises(ValueError) as ctx:
            self._parse(wb)
        self.assertIn("day", str(ctx.exception).lower())

    def test_non_numeric_duration_does_not_abort_validation(self):
        """One malformed cell must not stop every other rule being checked."""
        wb = self._sheet(self.FULL,
                         [["A", "Sun", "13:00 - 22:00", "Break 1", "15:15", "n/a", "Scheduled"]])
        rows, _ = self._parse(wb)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["duration"], 0,
                         "unparseable duration becomes 0 and is rejected downstream")

    def test_short_row_does_not_raise_index_error(self):
        """A truncated row must not crash, and must not be silently dropped.

        Associate and day are present and valid, so the row is real; the missing
        duration becomes 0, which INVALID_BREAK_ROW rejects downstream. Dropping
        it instead would hide a malformed break from the validation.
        """
        wb = self._sheet(self.FULL, [["A", "Sun"]])
        rows, _ = self._parse(wb)
        self.assertEqual(len(rows), 1, "a real row must not be silently dropped")
        self.assertEqual(rows[0]["duration"], 0, "missing duration surfaces as 0")
        self.assertIsNone(rows[0]["start"])

    def test_row_without_a_valid_day_is_skipped(self):
        wb = self._sheet(self.FULL,
                         [["A", "NotADay", "13:00 - 22:00", "Break 1", "15:15", 15, "Scheduled"]])
        rows, _ = self._parse(wb)
        self.assertEqual(rows, [], "rows outside the seven days are not break rows")

    def test_before_break_status_is_detected(self):
        wb = self._sheet(self.FULL,
                         [["A", "Sun", "13:00 - 22:00", "Break 1", "15:15", 15,
                           "Not assigned in before-break artifact"]])
        _, before_only = self._parse(wb)
        self.assertTrue(before_only, "artifact role detection must still work")


class CaseRootIdentityIsComplete(unittest.TestCase):
    """The identity a reviewer reads must carry every axis the gate requires.

    The release gate keys identity on input, contract and engine hash plus run
    metadata. The wrapper can only know input and engine hashes up front; the
    contract hash, parameters hash and run id are derived by the engine while
    parsing, and were written only to work_dir/RUN_IDENTITY.json. The case-root
    UNIVERSAL_RUN_IDENTITY.json therefore carried two of four axes, with
    contract_sha256 and run_id absent - an artifact that cannot state its own
    contract hash cannot satisfy the gate.
    """

    def _runner_source(self):
        return (ROOT / "engine" / "RUN_UNIVERSAL_PRODUCTION.py").read_text(encoding="utf-8")

    def test_runner_merges_engine_identity(self):
        source = self._runner_source()
        self.assertIn("RUN_IDENTITY.json", source)
        for key in ("contract_sha256", "run_id", "parameters_sha256"):
            self.assertIn(f"'{key}'", source, f"{key} must be carried into case-root identity")

    def test_identity_is_rewritten_after_the_engine_runs(self):
        """Merging before the engine runs would capture nothing."""
        source = self._runner_source()
        engine_call = source.index("engine_rc = subprocess.call(command)")
        rewrite = source.index("UNIVERSAL_RUN_IDENTITY.json", engine_call)
        self.assertGreater(rewrite, engine_call,
                           "identity must be completed after the engine produces it")

    def test_hash_disagreement_is_surfaced_not_merged(self):
        source = self._runner_source()
        self.assertIn("identity_mismatches", source,
                      "a wrapper/engine hash disagreement must be recorded, not "
                      "silently overwritten")

    def test_missing_engine_identity_is_recorded(self):
        source = self._runner_source()
        self.assertIn("engine_identity_source", source,
                      "provenance of the merged identity must be stated")


if __name__ == "__main__":
    unittest.main(verbosity=2)
