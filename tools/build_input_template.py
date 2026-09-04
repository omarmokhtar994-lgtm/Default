"""Rebuild an input workbook into the two-tab production template.

Sheet names stay 'Instructions' and 'Engine Defaults' because the engine
resolves those by alias; only their content and presentation change.

  Instructions   -> the ~20 decisions a scheduler makes per schedule, grouped,
                    with dropdowns where the value is an enum.
  Engine Defaults-> the engine tuning rows, defaults intact, marked do-not-touch.

Nothing is invented: every value is carried across from the source workbook.
Rows the engine does not read are dropped; rows it reads but the workbook never
set are added at their engine default so they are visible and changeable.
"""
import shutil, sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

YES_NO = '"Yes,No"'
SETUP_LAYOUT = [
    ("Case Setup", [
        ("Program Name", None),
        ("Count of Associates", None),
        ("Allow Headcount Mismatch", YES_NO),
        ("Interval Minutes", '"15,30,60"'),
        ("Requirements Source", '"FT Wise 15 Min,FT Wise 30 Min,FT Wise 60 Min"'),
        ("Shrinkage Source", '"Shrinkage 15 Min,Shrinkage 30 Min,Shrinkage 60 Min"'),
    ]),
    ("How To Run", [
        ("Run Stage", '"Before Breaks Only,Full Schedule"'),
        ("Run Depth", '"Quick,Deep,Overnight"'),
    ]),
    ("Coverage", [
        ("Target", None),
        ("Target Priority Confirmed", YES_NO),
        ("Minimum Per Interval", None),
        ("Hard Floor Solver Constraint Enabled", YES_NO),
        ("Blank Interval Staffing Rule", None),
    ]),
    ("Shift & OFF", [
        ("Allowed Shift Durations Hours", None),
        ("Use 11H/3OFF", YES_NO),
        ("Strict OFF Count", YES_NO),
        ("Separate OFF Days", YES_NO),
        ("Rest Gap Hours", None),
        ("Count of Different Shifts Per week", None),
    ]),
    ("Requests", [
        ("Fixed Request Use", YES_NO),
        ("Hard OFF Preferences", YES_NO),
        ("Leave Enabled", YES_NO),
        ("Use Preferences", YES_NO),
    ]),
    ("Opening Guard", [
        ("Opening Guard Enabled", YES_NO),
        ("Opening Minimum FTE", None),
        ("Opening Guard Intervals", None),
    ]),
    ("Breaks", [
        ("Short Break Count", None),
        ("Short Break Duration Minutes", None),
        ("Lunch Count", None),
        ("Lunch Duration Minutes", None),
        ("Break Preferred Gap Minutes", None),
        ("Break Absolute Minimum Gap Minutes", None),
        ("Break Normal Maximum Gap Minutes", None),
        ("Allow Back-to-Back Breaks", YES_NO),
    ]),
    ("Exception Policy", [
        ("Critical Coverage No-Break Exception Enabled", YES_NO),
        ("Critical Coverage No-Break Max Associate-Days", None),
    ]),
    ("Release Quality", [
        ("Production Quality Gate Mode", '"Warn,Fail,Off"'),
        ("Minimum After Break Target Ratio", None),
        ("Protected Before80 Minimum Intervals", None),
        ("Protected After80 Minimum Intervals", None),
    ]),
]
GATE_MODE = '"Warn,Fail,Off"'
ADVANCED_LAYOUT = [
    ("Demand Fit", [("Demand Fit Guard Enabled", '"Auto,Yes,No"'),
                    ("Demand Fit Minimum Active Minutes", None),
                    ("Demand Fit Minimum Active Ratio", None),
                    ("Demand Fit Maximum Blank Span Minutes", None)]),
    ("Overage Control", [("Overage Control Enabled", YES_NO), ("Overage Soft Cap", None),
                         ("Overage Severe Cap", None), ("Overage Extreme Cap", None),
                         ("Overage Penalty Weight", None)]),
    ("Break Concurrency", [("Maximum Concurrent Break Ratio", None),
                           ("Maximum Concurrent Breaks", None),
                           ("Break Concurrency Gate Mode", GATE_MODE)]),
    ("Week Boundary", [("Next Sunday Balance Enabled", YES_NO), ("Next Sunday Overage Cap", None),
                       ("Next Sunday Maximum Adjacent Raw Change", None),
                       ("Next Sunday Balance Gate Mode", GATE_MODE)]),
    ("Language / Skill", [("Language Operational Reserve Enabled", YES_NO),
                          ("Language Operational Reserve Extra FTE", None),
                          ("Language Reserve Gate Mode", GATE_MODE),
                          ("Qualified Language Break Certificate Enabled", YES_NO)]),
    ("Whole Week Balance", [("Whole Week Balance Enabled", YES_NO), ("Whole Week Overage Cap", None)]),
    ("Benchmarks", [("Quality Benchmark Tolerance Intervals", None)]),
]
DEAD_ROWS = {"rc9.1deepdefaultseconds", "rc9.1fulldefaultseconds",
             "rc9.1stage2searchorder", "rc9.1jointbudgetpolicy"}
# Rows the planner was asked to drive from the sheet. Leaving them blank is
# honest but useless: a dropdown you have to discover is not a choice you were
# offered. These two are seeded with the exact value the engine falls back to
# when the cell is empty, so pre-filling them changes no behaviour - it only
# makes the setting visible. Nothing else is seeded, because pre-filling a row
# whose default the engine may revise would freeze that default into the
# contract without anyone deciding to.
SEEDED_DEFAULTS = {"runstage": "Full Schedule", "rundepth": "Deep"}

HDR = PatternFill("solid", fgColor="1F3864")
SECTION = PatternFill("solid", fgColor="D9E2F3")
WARN = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

def norm(s): return "".join(ch for ch in str(s or "").lower() if ch.isalnum() or ch == ".")

def existing_values(wb):
    values = {}
    for name in ("Instructions", "Engine Defaults"):
        if name not in wb.sheetnames: continue
        for r in wb[name].iter_rows(values_only=True):
            if name == "Instructions":
                key, val = (r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
            else:
                key, val = (r[0], r[1] if len(r) > 1 else None)
            if key and norm(key) not in DEAD_ROWS:
                values.setdefault(norm(key), val)
    return values

def write_sheet(wb, title, layout, values, blurb, start_note=None):
    if title in wb.sheetnames: del wb[title]
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 62
    ws["A1"] = blurb
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = HDR
    ws.merge_cells("A1:D1")
    if start_note:
        ws["A2"] = start_note
        ws["A2"].fill = WARN
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells("A2:D2")
        ws.row_dimensions[2].height = 30
    row = 4 if start_note else 3
    for col, head in zip("ABCD", ("Section", "Instruction", "Value", "Notes")):
        c = ws[f"{col}{row}"]; c.value = head
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HDR
    row += 1
    validations = {}
    written = 0
    for section, items in layout:
        first = True
        for key, choices in items:
            v = values.get(norm(key))
            if v in (None, "") and norm(key) in SEEDED_DEFAULTS:
                v = SEEDED_DEFAULTS[norm(key)]
            ws.cell(row, 1, section if first else "").fill = SECTION if first else PatternFill()
            if first: ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 2, key)
            ws.cell(row, 3, v if v is not None else "")
            for col in range(1, 5): ws.cell(row, col).border = THIN
            if choices:
                dv = validations.get(choices)
                if dv is None:
                    dv = DataValidation(type="list", formula1=choices, allow_blank=True)
                    ws.add_data_validation(dv); validations[choices] = dv
                dv.add(ws.cell(row, 3))
            if v is None:
                ws.cell(row, 4, "not set - engine default applies")
                ws.cell(row, 4).font = Font(italic=True, color="808080")
            first = False
            row += 1; written += 1
        row += 1
    ws.freeze_panes = ws.cell(row=(5 if start_note else 4), column=1)
    return written

src, dst = Path(sys.argv[1]), Path(sys.argv[2])
shutil.copy(src, dst)
wb = load_workbook(dst)
values = existing_values(wb)
n1 = write_sheet(wb, "Instructions", SETUP_LAYOUT, values,
                 "SETUP - the decisions you make for this schedule",
                 "Fill the Value column. Cells with a dropdown accept only the listed options. "
                 "Anything left blank uses the engine default and is marked in Notes.")
n2 = write_sheet(wb, "Engine Defaults", ADVANCED_LAYOUT, values,
                 "ADVANCED - engine tuning. Defaults are production-tested.",
                 "Change these only with a specific reason. They are not per-schedule business "
                 "settings; the values here are the ones every released result was measured with.")
# Input Checks was a static snapshot that said PASS no matter what the roster held.
if "Input Checks" in wb.sheetnames:
    del wb["Input Checks"]
    ws = wb.create_sheet("Input Checks")
    ws["A1"] = "Input Checks are performed by the engine, not by this sheet."
    ws["A1"].font = Font(bold=True, color="FFFFFF"); ws["A1"].fill = HDR
    ws.merge_cells("A1:F1"); ws.column_dimensions["A"].width = 110
    ws["A3"] = ("The previous version of this tab held hardcoded PASS/WARN text. It reported "
                "PASS regardless of what the roster, demand or language setup actually "
                "contained, so it could not catch the errors it appeared to check for.")
    ws["A4"] = ("The engine runs the real pre-solver contract validation on every run and writes "
                "the result to the audit JSON as pre_solver_contract_validation, with a specific "
                "code, day and time for each failure. That is the authoritative check.")
    for r in (3, 4):
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 45
order = ["Instructions", "Engine Defaults"] + [s for s in wb.sheetnames if s not in ("Instructions", "Engine Defaults")]
wb._sheets = [wb[s] for s in order]
wb.save(dst)
print(f"{dst.name}: Setup {n1} rows, Advanced {n2} rows")
