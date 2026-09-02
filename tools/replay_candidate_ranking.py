#!/usr/bin/env python3
"""Replay a recorded candidate leaderboard under pre-fix and post-fix ranking.

Purpose
-------
The RC9.2.1 NMG EN evidence shows coverage improving while avoidable overage
rose 1103.08 -> 1244.11 FTE.  Two explanations are consistent with that:

  (a) the overage was *bought* - it is the honest price of +14 floor intervals; or
  (b) the overage was *avoidable* - a candidate with equal coverage and lower
      overage existed in the pool and lost because the deficit-masking defect
      terminated the comparison before the balance terms were consulted.

This tool distinguishes them without running a solver.  It reads the candidate
leaderboard the engine already exported during the run, re-ranks the identical
pool under both orderings, and reports whether the champion changes.

If the champion changes and the new champion has equal target/protected/floor
counts with lower overage, explanation (b) is confirmed and the defect
demonstrably cost the release.  If the champion is unchanged, explanation (a)
holds and the overage increase must be investigated elsewhere (aggregate
guidance, shift-library geometry, or break repair).

Usage
-----
    python tools/replay_candidate_ranking.py LEADERBOARD [--target 0.90] [--floor 0.75]

LEADERBOARD may be:
  * the exported .xlsx containing a candidate leaderboard sheet, or
  * a CSV with the same column headers.

Both orderings are computed from the engine's own selector, so this cannot
drift from production behaviour.
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "engine" / "_tools"))

import l632_universal_scheduler as E  # noqa: E402

# Leaderboard column header -> metrics key used by the selector.
COLUMN_MAP = {
    "Before Target": "before_target",
    "Before Floor": "before_floor",
    "Before100": "before_100",
    "Before90": "before_90",
    "Before80": "before_80",
    "After Target": "after_target",
    "After Floor": "after_floor",
    "After100": "after_100",
    "After90": "after_90",
    "After80": "after_80",
    "Severe Floor Gaps": "severe_floor_gap_count",
    "Max Floor Run": "max_consecutive_floor_gaps",
    "Floor Deficit Sum": "floor_deficit_sum",
    "Target Deficit Sum": "target_deficit_sum",
    "Before Avoidable Overage": "before_avoidable_overage_fte_sum",
    "After Avoidable Overage": "after_avoidable_overage_fte_sum",
    "Before Target Overage": "before_target_overage_fte_sum",
    "After Target Overage": "after_target_overage_fte_sum",
    "Before Severe Overage": "before_severe_overage_count",
    "After Severe Overage": "after_severe_overage_count",
    "Before Extreme Overage": "before_extreme_overage_count",
    "After Extreme Overage": "after_extreme_overage_count",
}


def _number(value):
    if value is None or value == "":
        return 0
    try:
        f = float(value)
    except (TypeError, ValueError):
        return value
    return int(f) if f.is_integer() else f


def load_rows(path: Path):
    """Return (label, metrics-dict) pairs from an xlsx or csv leaderboard."""
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            header = [str(c.value).strip() if c.value is not None else ""
                      for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "Floor Deficit Sum" not in header:
                continue
            for raw in ws.iter_rows(min_row=2, values_only=True):
                if raw is None or all(v is None for v in raw):
                    continue
                record = dict(zip(header, raw))
                yield _to_metrics(record)
            return
        raise SystemExit(
            f"No candidate leaderboard sheet found in {path}. "
            "Expected a sheet with a 'Floor Deficit Sum' column."
        )
    with path.open(newline="", encoding="utf-8-sig") as fh:
        for record in csv.DictReader(fh):
            yield _to_metrics(record)


def _to_metrics(record: dict):
    label = " / ".join(
        str(record.get(k)) for k in ("Skeleton Profile", "Break Profile", "Objective Mode")
        if record.get(k)
    ) or "candidate"
    metrics = {}
    for column, key in COLUMN_MAP.items():
        if column in record:
            metrics[key] = _number(record[column])
    # Mirror deficits into the prefixed keys the fixed selector prefers, and
    # into the before_* severe/run keys the before-prefix path reads.
    for prefix in ("before", "after"):
        metrics.setdefault(f"{prefix}_floor_deficit_sum", metrics.get("floor_deficit_sum", 0))
        metrics.setdefault(f"{prefix}_target_deficit_sum", metrics.get("target_deficit_sum", 0))
        metrics.setdefault(f"{prefix}_floor_deficit_max", metrics.get("floor_deficit_max", 0))
    metrics.setdefault("before_severe_floor_gap_count", metrics.get("severe_floor_gap_count", 0))
    metrics.setdefault("before_max_consecutive_floor_gaps",
                       metrics.get("max_consecutive_floor_gaps", 0))
    metrics.setdefault("selected_role", record.get("Selected Role", ""))
    return label, metrics


def deficit_term_index(parsed, metrics, prefix) -> int:
    """Locate the quantized floor-deficit terms inside the quality tuple.

    The two deficit terms do NOT sit at a fixed offset.  `_candidate_quality_tuple`
    splats `_protected_tier_counts` before them, and that helper returns a
    variable number of tiers depending on the contract:

        target 90%, floor 75%  -> 1 tier  (80)      -> deficit terms at 9, 10
        target 100%, floor 75% -> 2 tiers (90, 80)  -> deficit terms at 10, 11
        target 80%, floor 75%  -> 0 tiers           -> deficit terms at 8, 9

    Hardcoding 9/10 was correct only for the 90/75 contract this tool was first
    used on.  Under 100/75 it overwrote `-max_consecutive_floor_gaps` - a real
    safety term - with a raw deficit float, and under 80/75 it overwrote both
    gap terms, so the "pre-fix" ranking was not the pre-fix ranking at all and
    the verdict was silently wrong.  The index is now derived and then verified
    against the tuple's actual contents, so a future change to the tuple's shape
    fails loudly here instead of producing bogus release evidence.
    """
    index = 8 + len(E._protected_tier_counts(parsed, metrics, prefix))
    tup = E._candidate_quality_tuple(parsed, metrics, prefix)
    expected = (
        -E._deficit_bucket(E._prefixed_metric(metrics, prefix, "floor_deficit_sum")),
        -E._deficit_bucket(E._prefixed_metric(metrics, prefix, "floor_deficit_max")),
    )
    if index + 1 >= len(tup) or (tup[index], tup[index + 1]) != expected:
        raise SystemExit(
            "Cannot locate the quantized deficit terms in the quality tuple "
            f"(derived index {index}, found {tup[index:index + 2]}, "
            f"expected {expected}). _candidate_quality_tuple has changed shape; "
            "update deficit_term_index before trusting this replay.")
    return index


def prefix_ranking(parsed, metrics, prefix, masked: bool, index: int):
    """Fixed ordering, or the pre-fix ordering with raw floats restored."""
    tup = list(E._candidate_quality_tuple(parsed, metrics, prefix))
    if masked:
        tup[index] = -float(metrics.get(f"{prefix}_floor_deficit_sum",
                                        metrics.get("floor_deficit_sum", 0.0)) or 0.0)
        tup[index + 1] = -float(metrics.get(f"{prefix}_floor_deficit_max",
                                            metrics.get("floor_deficit_max", 0.0)) or 0.0)
    return tuple(tup)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("leaderboard", type=Path)
    ap.add_argument("--target", type=float, default=0.90,
                    help="Workbook target ratio (NMG EN = 0.90)")
    ap.add_argument("--floor", type=float, default=0.75,
                    help="Configured floor ratio (NMG EN = 0.75)")
    ap.add_argument("--active", type=int, default=252,
                    help="Active interval count (NMG EN = 252)")
    ap.add_argument("--prefix", choices=["before", "after"], default="before",
                    help="Rank before-break skeletons or after-break candidates")
    args = ap.parse_args()

    if not args.leaderboard.exists():
        raise SystemExit(f"Leaderboard not found: {args.leaderboard}")

    parsed = SimpleNamespace(target_ratio=args.target, floor_ratio=args.floor,
                             active=[[True] * args.active])
    pool = list(load_rows(args.leaderboard))
    for _, metrics in pool:
        metrics.setdefault("active_intervals", args.active)
    if not pool:
        raise SystemExit("Leaderboard contained no candidate rows.")

    # Derived once from the contract, then verified against every row: rows can
    # carry different metric keys, and a row whose tuple has a different shape
    # must not be ranked on the wrong terms.
    index = deficit_term_index(parsed, pool[0][1], args.prefix)
    for _, metrics in pool:
        deficit_term_index(parsed, metrics, args.prefix)

    fixed = max(pool, key=lambda p: prefix_ranking(parsed, p[1], args.prefix, False, index))
    masked = max(pool, key=lambda p: prefix_ranking(parsed, p[1], args.prefix, True, index))

    ov = f"{args.prefix}_avoidable_overage_fte_sum"
    tgt, flr = f"{args.prefix}_target", f"{args.prefix}_floor"

    print(f"candidates in pool        : {len(pool)}")
    print(f"contract                  : target {args.target:.0%}, floor {args.floor:.0%}, "
          f"{args.active} active intervals")
    print(f"deficit quantum           : {E.FLOOR_DEFICIT_COMPARISON_QUANTUM}")
    print()
    print(f"pre-fix champion          : {masked[0]}")
    print(f"    target/floor          : {masked[1].get(tgt)} / {masked[1].get(flr)}")
    print(f"    avoidable overage FTE : {masked[1].get(ov)}")
    print(f"post-fix champion         : {fixed[0]}")
    print(f"    target/floor          : {fixed[1].get(tgt)} / {fixed[1].get(flr)}")
    print(f"    avoidable overage FTE : {fixed[1].get(ov)}")
    print()

    if masked[0] == fixed[0]:
        print("VERDICT: champion UNCHANGED.")
        print("  The deficit-masking defect did not cost this run. The overage")
        print("  increase was bought by the coverage gain, or originates outside")
        print("  candidate selection (aggregate guidance, shift geometry, break")
        print("  repair). Investigate those next.")
        return 0

    delta = float(masked[1].get(ov, 0) or 0) - float(fixed[1].get(ov, 0) or 0)
    same_coverage = (masked[1].get(tgt) == fixed[1].get(tgt)
                     and masked[1].get(flr) == fixed[1].get(flr))
    print("VERDICT: champion CHANGED.")
    print(f"  Avoidable overage difference: {delta:+.2f} FTE "
          f"({'saved by the fix' if delta > 0 else 'worse under the fix'})")
    if same_coverage and delta > 0:
        print("  Target and floor counts are IDENTICAL between the two champions,")
        print("  so this overage was avoidable at no coverage cost. The masking")
        print("  defect demonstrably cost the release. Root cause CONFIRMED.")
    elif not same_coverage:
        print("  Coverage counts DIFFER between champions - inspect the trade")
        print("  before claiming a win; the priority model must still hold.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
