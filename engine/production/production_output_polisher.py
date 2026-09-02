#!/usr/bin/env python3
"""Artifact publisher for the Phase C C4 phase-closure master engine development checkpoint."""
from __future__ import annotations
import argparse,csv,hashlib,json,re,shutil,tempfile,zipfile
from datetime import datetime,timezone
from pathlib import Path

# Fallbacks only. These literals are what the published artifact used to assert
# unconditionally, and they had drifted badly: RELEASE named RC9.1 while the
# engine being published was RC9.2.1, and ENGINE was a hash matching no engine
# in the project - not RC9.1's da21c3ba, not RC9.2.1's 56ec2eef, not the current
# build. Every published workbook, release manifest and production ZIP filename
# therefore asserted a fabricated engine identity, which is precisely what the
# exact-engine-identity release gate exists to prevent.
#
# Identity is now derived from the run that produced the artifact. These remain
# only so a case root without an identity file still publishes, and when they
# are used the manifest records identity_source='FALLBACK_LITERALS' so the
# substitution is visible rather than silent.
RELEASE='L6.3.2.3-RC9.1-UNIVERSAL-SEARCH-RECOVERY-PLATFORM'
SOLVER='L6.3.2.3-RC9.1-UNIVERSAL-SEARCH-RECOVERY-PLATFORM'
COMMIT='RC9_UNIVERSAL_PLATFORM_BASED_ON_RC8_13_TARGET90_CORE'
ENGINE='4b5f58e5df5513057f511d9e13ccece45226ce548aaa2f5f9a3ecd13fc6eb31d'
APPROVAL='WORKBOOK_DRIVEN_RELEASE_STATUS_NOT_CASE_PRESET'

def run_identity(root):
 '''Identity of the run that produced this case, from its own artifacts.'''
 ident={'release':RELEASE,'solver':SOLVER,'commit':COMMIT,'engine_sha256':ENGINE,
        'contract_sha256':None,'run_id':None,'input_sha256':None,
        'identity_source':'FALLBACK_LITERALS'}
 # Read BOTH sources and fill gaps, rather than stopping at the first. A case
 # produced before the wrapper learned to merge the engine's identity has a
 # UNIVERSAL_RUN_IDENTITY.json without contract_sha256 or run_id, while
 # debug/RUN_IDENTITY.json alongside it carries them. Stopping at the first file
 # would publish those older cases with the contract hash still missing.
 sources=[]
 for name in ('UNIVERSAL_RUN_IDENTITY.json','debug/RUN_IDENTITY.json'):
  p=root/name
  if not p.is_file(): continue
  try: d=json.loads(p.read_text(encoding='utf-8'))
  except (OSError,json.JSONDecodeError): continue
  release=d.get('release') or d.get('version')
  if release and ident['identity_source']=='FALLBACK_LITERALS':
   ident['release']=release; ident['solver']=release
  for src,dst in (('engine_sha256','engine_sha256'),('contract_sha256','contract_sha256'),
                  ('run_id','run_id'),('input_sha256','input_sha256'),('git_commit','commit')):
   if d.get(src) and ident.get(dst) in (None,ENGINE,COMMIT): ident[dst]=d[src]
  sources.append(name)
  if ident['identity_source']=='FALLBACK_LITERALS': ident['identity_source']=name
 if len(sources)>1: ident['identity_source']='+'.join(sources)
 return ident
ORDER=['Production Summary','Schedule','Break Schedule','Break Spacing Audit','FT Wise After Breaks','Coverage Before Breaks','Interval Coverage Audit','Overage Audit','Next Sunday Carry-Out Audit','Canonical Contract','No-Break Exceptions','Validation Log','Rule Checks','Rest Gap Audit','Language Skill Audit','Language Reserve Summary','Skill Allocation Audit','Whole Week Balance Audit','Employee Quality Audit','Feasibility Certificate','Language Setup','Preference','Instructions','Fixed Shift Requests','Candidate Leaderboard','Target Tradeoff Audit','Feasibility Report','Blank Interval Audit','Shift Demand Fit Audit','Overnight Audit','Cyclic Sunday Audit']
TECH={'Final Schedule','Break Schedule Active','Daily Interval Review','FT Wise Active','FT Wise After Breaks Active','Scheduler Engine','Previous Engine','Balanced Scenario Schedule','Future Use','Implementation Notes','Formula Fix Notes','Review Runs','Balance Change Log','Benchmark Comparison','Dynamic Interval Guide','Constraint Isolation','Day Tail Fit Audit'}

def sha(p):
 h=hashlib.sha256()
 with open(p,'rb') as f:
  for c in iter(lambda:f.read(1048576),b''): h.update(c)
 return h.hexdigest()
def one(root,pat):
 m=sorted(root.glob(pat))
 if not m: raise FileNotFoundError(f'{pat} in {root}')
 return m[0]
def one_of(root,*patterns):
 for pat in patterns:
  m=sorted(root.glob(pat))
  if m: return m[0]
 raise FileNotFoundError(f'{patterns} in {root}')
def read_csv(p):
 with p.open(newline='',encoding='utf-8') as f: r=list(csv.DictReader(f))
 return r[0] if r else {}
def mkey(m): return tuple(m.get(k) for k in ('after_100','after_90','after_80','after_target','after_floor','hard_floor_gap_count','week_boundary_after_target','week_boundary_after_floor','week_boundary_floor_gap_count','week_boundary_zero_staffed_active_quarters','week_boundary_language_gap_count','week_boundary_opening_gap_count','week_boundary_blank_staffed_quarters','severe_floor_gap_count','max_consecutive_floor_gaps','floor_deficit_sum'))
def roles(pareto):
 ex=pareto.get('exports') or []; rec=next((x for x in ex if x.get('role') in {'RECOMMENDED_FINAL','BEST_FINAL_AFTER_BREAKS'}),{}); rk=mkey(rec.get('metrics') or {}); out=[]
 for x in ex:
  role=x.get('role',''); m=x.get('metrics') or {}
  if role=='BEST_BEFORE_BREAKS': disp='Published as BEST_BEFORE_BREAKS_SCHEDULE (review only)'
  elif role in {'RECOMMENDED_FINAL','BEST_FINAL_AFTER_BREAKS'}: disp='Published as BEST_FINAL_AFTER_BREAKS_SCHEDULE'
  elif mkey(m)==rk: disp='Shared with BEST_FINAL_AFTER_BREAKS_SCHEDULE; duplicate workbook suppressed'
  else: disp='Distinct alternative retained in leaderboard/Pareto/debug evidence only'
  out.append([role,m.get('after_100',''),m.get('after_90',''),m.get('after_80',''),m.get('after_target',''),m.get('after_floor',''),disp])
 return out
def clean_book(src,dst,role,use,pareto,ident):
 from openpyxl import load_workbook
 from openpyxl.styles import Alignment,Font,PatternFill
 from openpyxl.workbook.properties import CalcProperties
 dst.parent.mkdir(parents=True,exist_ok=True)
 with tempfile.TemporaryDirectory() as td:
  stage=Path(td)/dst.name; shutil.copy2(src,stage); wb=load_workbook(stage)
  if 'Prefrence' in wb.sheetnames and 'Preference' not in wb.sheetnames: wb['Prefrence'].title='Preference'
  for ws in wb.worksheets:
   for row in ws.iter_rows():
    for cell in row:
     if isinstance(cell.value,str):
      cell.value=cell.value.replace('Dynamic Schedule - V10 interval-aware optimized output','Universal WFM Production Schedule').replace('V10 Dynamic Interval Pattern MILP optimizer','Universal WFM CP-SAT Scheduler')
      cell.value=re.sub(r'\bPrefrence\b','Preference',cell.value,flags=re.I)
  if 'Production Summary' not in wb.sheetnames:
   ps=wb.create_sheet('Production Summary',0); ps.append(['Metric','Value'])
  ps=wb['Production Summary']; existing={str(ps.cell(r,1).value or '').strip():r for r in range(1,ps.max_row+1)}
  fields={'Artifact Type':role,'Artifact Role':role,'Production Use':use,'Production Release':ident['release'],'Production Approval':APPROVAL,'Solver Version':ident['solver'],'Week-Boundary Patch Commit':ident['commit'],'Engine SHA256':ident['engine_sha256'],'Contract SHA256':ident['contract_sha256'] or 'NOT_RECORDED','Run ID':ident['run_id'] or 'NOT_RECORDED','Identity Source':ident['identity_source'],'Validation Basis':'RC9 is a universal workbook-driven production platform. Artifact verification, hard validity, target/floor quality, and any quality-debt approval are reported separately for each uploaded workbook.','Week-Boundary Protection':'Current Saturday carry-out and next-Sunday active intervals are included in final release gates','Approved Waiver':'None encoded by program/client name. Waivers must come from workbook contract settings or documented human approval after the run.','Known Limitation':'No client-specific release pass is implied. Each uploaded workbook is solved and validated from its own contract; quality-blocked schedules require explicit approval before operational use.'}
  for k,v in fields.items():
   if k in existing: ps.cell(existing[k],2).value=v
   else: ps.append([k,v])
  ps.append([]); ps.append(['CANDIDATE ROLE CONSOLIDATION','After100','After90','After80','After Target','After Floor','Production Disposition'])
  for row in roles(pareto): ps.append(row)
  ps.sheet_view.showGridLines=False; ps.freeze_panes='A2'; ps.column_dimensions['A'].width=38; ps.column_dimensions['B'].width=105; ps.column_dimensions['G'].width=72
  for cell in ps[1]: cell.fill=PatternFill('solid',fgColor='1F4E78'); cell.font=Font(color='FFFFFF',bold=True); cell.alignment=Alignment(horizontal='center')
  by={ws.title:ws for ws in wb.worksheets}; ordered=[]; used=set()
  for name in ORDER:
   if name in by: ordered.append(by[name]); used.add(name)
  ordered += [ws for ws in wb.worksheets if ws.title not in used]; wb._sheets=ordered
  for ws in wb.worksheets:
   ws.sheet_state='hidden' if ws.title in TECH else 'visible'
  wb.active=wb.sheetnames.index('Production Summary')
  if getattr(wb,'calculation',None) is None: wb.calculation=CalcProperties(calcMode='auto')
  wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.save(dst); wb.close()
 with zipfile.ZipFile(dst) as z:
  bad=z.testzip()
  if bad: raise RuntimeError(bad)
 wb=load_workbook(dst,read_only=True); names=wb.sheetnames; wb.close()
 if not {'Production Summary','Schedule','Canonical Contract'}.issubset(names): raise RuntimeError('Required sheet missing')
 return {'path':str(dst),'size_bytes':dst.stat().st_size,'sha256':sha(dst),'sheet_count':len(names),'xlsx_zip_test':'PASS','reopen_test':'PASS'}
def publish(root):
 root=root.resolve(); summary=read_csv(one(root,'*.l6_3_2_3_summary.csv')); pareto=json.loads(one(root,'*_PARETO_EXPORT_MANIFEST.json').read_text()); cid=Path(summary.get('output','')).name.split('_L6_3_2_3_',1)[0]; prod=root/'production'; prod.mkdir(exist_ok=True); ident=run_identity(root)
 before=clean_book(one(root,'*_BEST_BEFORE_BREAKS_SCHEDULE.xlsx'),prod/f'{cid}_L6_3_2_3_BEST_BEFORE_BREAKS_SCHEDULE.xlsx','BEST_BEFORE_BREAKS_SCHEDULE','REVIEW ONLY - strongest legal shift/OFF skeleton before breaks; not operational',pareto,ident)
 final=clean_book(one_of(root,'*_BEST_FINAL_AFTER_BREAKS_SCHEDULE.xlsx','*_RECOMMENDED_FINAL_AFTER_BREAKS_SCHEDULE.xlsx'),prod/f'{cid}_L6_3_2_3_BEST_FINAL_AFTER_BREAKS_SCHEDULE.xlsx','BEST_FINAL_AFTER_BREAKS_SCHEDULE','PRODUCTION OUTPUT - use only when hard gates pass or a documented quality-debt approval exists',pareto,ident)
 manifest={'release':ident['release'],'identity_source':ident['identity_source'],'contract_sha256':ident['contract_sha256'],'run_id':ident['run_id'],'input_sha256':ident['input_sha256'],'approval_status':APPROVAL,'generated_utc':datetime.now(timezone.utc).isoformat(),'case':cid,'solver':{'version':ident['solver'],'week_boundary_patch_commit':ident['commit'],'engine_sha256':ident['engine_sha256'],'optimization_logic_changed':True,'change_scope':'RC9 preserves workbook target-first skeleton/final selection, protected target-champion break proof, joint/break refinement, and workbook-driven quality gates without client/case branching'},'two_artifact_contract':{'BEST_BEFORE_BREAKS_SCHEDULE':before,'BEST_FINAL_AFTER_BREAKS_SCHEDULE':final},'candidate_role_rows':roles(pareto),'source_status':summary.get('status','')}
 mp=prod/'PRODUCTION_ARTIFACT_MANIFEST.json'; mp.write_text(json.dumps(manifest,indent=2),encoding='utf-8')
 ev=[]
 for pat in ('*.l6_3_2_3_summary.csv','*.l6_3_2_3_solver_audit.json','*_CANDIDATE_LEADERBOARD.csv','*_PARETO_EXPORT_MANIFEST.json','PHASE_C_QUALITY_SUMMARY.json','PHASE_C_QUALITY_SUMMARY.csv','debug/RUN_IDENTITY.json','debug/scheduler.log'):
  ev += [p for p in root.glob(pat) if p.is_file()]
 zp=prod/f'{cid}_{ident["release"]}_PRODUCTION_ONLY.zip'
 with zipfile.ZipFile(zp,'w',zipfile.ZIP_DEFLATED) as z:
  z.write(Path(before['path']),arcname=Path(before['path']).name); z.write(Path(final['path']),arcname=Path(final['path']).name); z.write(mp,arcname=mp.name)
  for p in sorted(set(ev)): z.write(p,arcname=f'evidence/{p.name}')
 with zipfile.ZipFile(zp) as z:
  bad=z.testzip()
  if bad: raise RuntimeError(bad)
 print(json.dumps({'status':'PASS','case':cid,'production_zip':str(zp),'before':before,'final':final},indent=2))
def main():
 ap=argparse.ArgumentParser(); ap.add_argument('--case-root',type=Path); ap.add_argument('--selfcheck',action='store_true'); a=ap.parse_args()
 if a.selfcheck:
  # Not `assert`: under `python -O` assertions are stripped and the selfcheck
  # would print PASS having verified nothing.
  if 'duplicate workbook suppressed' not in roles({'exports':[{'role':'RECOMMENDED_FINAL','metrics':{'after_100':1}},{'role':'MAX_FLOOR_CANDIDATE','metrics':{'after_100':1}}]})[1][-1]:
   print('PRODUCTION OUTPUT POLISHER SELFCHECK: FAIL (duplicate-role consolidation)'); return 1
  fb=run_identity(Path(tempfile.gettempdir())/'__rc921_no_such_case_root__')
  if fb['identity_source']!='FALLBACK_LITERALS':
   print('PRODUCTION OUTPUT POLISHER SELFCHECK: FAIL (identity fallback not marked)'); return 1
  print('PRODUCTION OUTPUT POLISHER SELFCHECK: PASS'); return 0
 if not a.case_root: raise SystemExit('--case-root required')
 publish(a.case_root); return 0
if __name__=='__main__': raise SystemExit(main())
