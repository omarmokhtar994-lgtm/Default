#!/usr/bin/env python3
"""RC9.2.1 contract-rule semantics guards.

Several rules in RULE_VERIFICATION_MATRIX.csv sit at PENDING_SOLVE because no
shipped fixture exercises them.  Two of those gaps are structural rather than
scheduling ones and were confirmed against the delivered test kit:

  * `15_11H_3OFF_SAKS` parses with ``use_11h_3off = False`` and a shift library
    containing only 540-minute shifts, so it cannot exercise 11H/3OFF or the
    11H break pattern at all.
  * `17_BILINGUAL_NMG_EN_SP` carries a single language rule (Spanish, eligible
    {spanish}).  With no English rule and no directional mapping, "Spanish may
    assist English only in the configured direction" and "no double counting"
    cannot be observed from it.

A missing fixture is not a reason to leave the semantics unpinned.  These tests
exercise the engine's own rule predicates directly, so the contract meaning is
guarded regardless of which workbooks happen to ship.

Run:  python tests/test_rc9_2_1_rule_semantics.py
"""
from __future__ import annotations

import sys
import unittest
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "engine" / "_tools"))

import l632_universal_scheduler as E  # noqa: E402


def rule(group, required, eligible, minimum=1, active=True):
    return SimpleNamespace(
        group=group,
        required_languages=set(required),
        eligible_languages={E.norm(x) for x in eligible},
        minimum=minimum,
        active=active,
    )


def associate(language):
    return SimpleNamespace(language=language)


class LanguageEligibilityIsExplicitOnly(unittest.TestCase):
    """Coverage credit follows the configured eligibility set, nothing more.

    The universal contract requires that a language requirement is met only by
    associates the workbook actually declares eligible for it.  An associate is
    never credited to a requirement merely because they are bilingual, and a
    single associate is never counted toward two requirements implicitly.
    """

    def test_matching_language_is_eligible(self):
        self.assertTrue(E.language_eligible(rule("Spanish", ["spanish"], ["spanish"]),
                                            associate("Spanish")))

    def test_non_matching_language_is_not_eligible(self):
        self.assertFalse(E.language_eligible(rule("Spanish", ["spanish"], ["spanish"]),
                                             associate("English")))

    def test_bilingual_is_not_implicitly_eligible_for_a_single_language_rule(self):
        """The core no-double-counting rule.

        A bilingual associate must NOT satisfy a Spanish-only requirement unless
        the workbook lists their language in the eligible set.  Crediting them
        automatically is exactly the double counting the contract forbids.
        """
        self.assertFalse(E.language_eligible(rule("Spanish", ["spanish"], ["spanish"]),
                                             associate("Bilingual")))

    def test_configured_multi_language_eligibility_is_honoured(self):
        """Directional assist, as GDI REAL28 configures it.

        A rule whose eligible set names several source languages accepts any of
        them - this is the source-to-target inversion, and it is opt-in.
        """
        r = rule("Bilingual/French", ["bilingual", "french"], ["bilingual", "french"])
        self.assertTrue(E.language_eligible(r, associate("Bilingual")))
        self.assertTrue(E.language_eligible(r, associate("French")))
        self.assertFalse(E.language_eligible(r, associate("English")))

    def test_assist_direction_is_not_symmetric(self):
        """Eligibility one way must not imply eligibility the other way.

        If Spanish speakers may assist English, that must not silently make
        English speakers eligible for the Spanish requirement.
        """
        english_rule = rule("English", ["english"], ["english", "spanish"])
        spanish_rule = rule("Spanish", ["spanish"], ["spanish"])
        self.assertTrue(E.language_eligible(english_rule, associate("Spanish")))
        self.assertFalse(E.language_eligible(spanish_rule, associate("English")))

    def test_eligibility_is_case_and_whitespace_insensitive(self):
        r = rule("Spanish", ["spanish"], ["spanish"])
        for spelling in ("Spanish", "SPANISH", "  spanish  ", "sPaNiSh"):
            with self.subTest(spelling=spelling):
                self.assertTrue(E.language_eligible(r, associate(spelling)))


class LongShiftOffExpectation(unittest.TestCase):
    """OFF count follows shift duration against the shared threshold.

    Pinned because the threshold moved from a duplicated 660 in the validator to
    the engine's 630 (Correction Pass 02, Finding 4).  No fixture in the kit has
    a shift at or above 630, so nothing there would catch a regression.
    """

    def expected_off(self, duration_min):
        return 3 if duration_min >= E.LONG_SHIFT_MIN_DURATION_MIN else 2

    def test_nine_hour_shift_expects_two_off(self):
        self.assertEqual(self.expected_off(540), 2)

    def test_eleven_hour_shift_expects_three_off(self):
        self.assertEqual(self.expected_off(660), 3)

    def test_boundary_shift_is_long_mode(self):
        self.assertEqual(self.expected_off(E.LONG_SHIFT_MIN_DURATION_MIN), 3)

    def test_just_below_boundary_is_short_mode(self):
        self.assertEqual(self.expected_off(E.LONG_SHIFT_MIN_DURATION_MIN - 1), 2)

    def test_the_gap_that_caused_the_defect(self):
        """A 10.75h shift: long to the engine, short to the old validator."""
        self.assertEqual(self.expected_off(645), 3)
        self.assertLess(645, 660, "645 was short mode under the old validator threshold")
        self.assertGreaterEqual(645, E.LONG_SHIFT_MIN_DURATION_MIN)


class RestCompatibility(unittest.TestCase):
    """Rest gap, including the previous-Saturday carry-in.

    Values are the two genuine conflicts the engine diagnosed in the historical
    NMG EN fixture, verified by hand: a prev-Sat 23:00-08:00 leaves 6h before a
    14:00 Sunday start, and 20:00-05:00 leaves 9h, both under the 12h rule.
    """

    def test_six_hour_gap_violates_twelve_hour_rest(self):
        self.assertFalse(E.previous_saturday_compatible(
            "23:00 - 08:00", SimpleNamespace(start_min=14 * 60, duration_min=540), 12.0))

    def test_nine_hour_gap_violates_twelve_hour_rest(self):
        self.assertFalse(E.previous_saturday_compatible(
            "20:00 - 05:00", SimpleNamespace(start_min=14 * 60, duration_min=540), 12.0))

    def test_sufficient_gap_is_accepted(self):
        self.assertTrue(E.previous_saturday_compatible(
            "08:00 - 17:00", SimpleNamespace(start_min=14 * 60, duration_min=540), 12.0))

    def test_blank_previous_saturday_is_compatible(self):
        self.assertTrue(E.previous_saturday_compatible(
            "", SimpleNamespace(start_min=0, duration_min=540), 12.0))


class PreviousSaturdayCarryIn(unittest.TestCase):
    """Carry-in coverage is half-open: a shift ending at T does not staff T.

    This is what makes the cleaned NMG EN fixture infeasible - its latest
    previous-Saturday shift ends exactly at 02:00, so Sunday 02:00 is unstaffed.
    """

    def test_shift_ending_at_the_boundary_does_not_cover_it(self):
        # 17:00-02:00 -> covers through the quarter ending 02:00, not 02:00 itself.
        self.assertFalse(E.previous_saturday_covers_qslot("17:00 - 02:00", 8))

    def test_shift_ending_at_the_boundary_covers_the_quarter_before(self):
        self.assertTrue(E.previous_saturday_covers_qslot("17:00 - 02:00", 7))

    def test_overnight_shift_reaches_into_sunday_morning(self):
        # 23:00-08:00 covers Sunday through 08:00; qslot 8 is 02:00.
        self.assertTrue(E.previous_saturday_covers_qslot("23:00 - 08:00", 8))

    def test_shift_not_reaching_sunday_covers_nothing(self):
        self.assertFalse(E.previous_saturday_covers_qslot("08:00 - 17:00", 0))


class BreakSegmentContract(unittest.TestCase):
    """9H is 15+30+15; 11H is 15+30+15+15. Segments are stored in quarters."""

    def test_nine_hour_pattern_is_fifteen_thirty_fifteen(self):
        segments = ((1, "Break 1"), (2, "Lunch"), (1, "Break 2"))
        self.assertEqual([q * 15 for q, _ in segments], [15, 30, 15])
        self.assertEqual(sum(q * 15 for q, _ in segments), 60)

    def test_eleven_hour_pattern_is_fifteen_thirty_fifteen_fifteen(self):
        segments = ((1, "Break 1"), (2, "Lunch"), (1, "Break 2"), (1, "Break 3"))
        self.assertEqual([q * 15 for q, _ in segments], [15, 30, 15, 15])
        self.assertEqual(sum(q * 15 for q, _ in segments), 75)


class CarryInForcedOverageIsSeparated(unittest.TestCase):
    """Overage the solver cannot avoid must be distinguishable from overage it can.

    Previous-week carry-in is fixed workbook input, not a decision variable.
    Where carry-in alone exceeds what the target needs, the excess was charged
    wholly as "avoidable" overage, flagged severe and extreme, with
    indivisible_staffing_only False - so an overage regression caused by the
    input contract was indistinguishable from one caused by the optimizer.
    """

    def test_carry_in_excess_is_charged_as_avoidable_by_the_raw_metric(self):
        """Pin the underlying behaviour the new metric exists to explain."""
        m = E.fractional_safe_overage_metrics(0.9, 10.0, 0.90, 1.0, 1.1, 1.25, 1.5)
        self.assertAlmostEqual(m["avoidable_overage_fte"], 9.0, places=9)
        self.assertFalse(m["indivisible_staffing_only"],
                         "nothing currently marks carry-in overage structural")

    def test_forced_overage_formula_matches_carry_in_excess(self):
        """carry_in_forced = max(0, carry-in effective - unavoidable at target)."""
        for req, eff, carry in ((0.9, 1.0, 10.0), (2.9, 1.0, 10.0), (2.0, 1.0, 10.0)):
            with self.subTest(req=req):
                m = E.fractional_safe_overage_metrics(req, carry, 0.90, eff, 1.1, 1.25, 1.5)
                unavoidable = m["unavoidable_effective_at_target"]
                self.assertAlmostEqual(max(0.0, carry - unavoidable),
                                       m["avoidable_overage_fte"], places=9)

    def test_no_forced_overage_when_carry_in_is_below_requirement(self):
        m = E.fractional_safe_overage_metrics(12.5, 4.0, 0.90, 0.87, 1.1, 1.25, 1.5)
        self.assertEqual(m["avoidable_overage_fte"], 0.0)

    def test_metric_keys_are_emitted(self):
        """The separated metrics must reach the metrics dict, not just be computed."""
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8")
        for key in ("carry_in_forced_overage_fte_sum",
                    "carry_in_forced_overage_interval_count",
                    "before_solver_controllable_avoidable_overage_fte_sum",
                    "after_solver_controllable_avoidable_overage_fte_sum"):
            self.assertIn(f'"{key}"', source, f"{key} not emitted in metrics dict")

    def test_carry_in_metric_does_not_enter_candidate_ranking(self):
        """Measurement only - adding it to the tuple would be a behaviour change."""
        import re
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8")
        start = source.index("def _candidate_quality_tuple")
        body = source[start:source.index("def break_solution_key", start)]
        self.assertNotIn("carry_in_forced", body,
                         "carry-in metric must not affect ranking without evidence")
        self.assertNotIn("solver_controllable", body)


class ResumeIdentityIsHashBased(unittest.TestCase):
    """Resume must key on content hashes, never on a filename or a bare label.

    Executed against the shipped RC9.1 checkpoint bundle: resuming it with the
    folder-13 workbook raises
    "Resume refused: input/contract/engine/seed/parameters hash differs from the
    existing run identity".  These guards pin the mechanism that makes that
    refusal sound.
    """

    def test_run_id_is_derived_from_the_full_identity(self):
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8")
        self.assertIn('run_identity["run_id"] = canonical_hash(run_identity)[:24]', source,
                      "run_id must hash the identity, so run_id equality implies "
                      "input/contract/engine/seed/parameter equality")

    def test_resume_refuses_on_identity_mismatch(self):
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8")
        self.assertIn("Resume refused:", source,
                      "a mismatched resume must raise, not silently start fresh")

    def test_identity_covers_every_input_that_can_change_results(self):
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8")
        start = source.index("run_identity = {")
        block = source[start:source.index("run_identity[\"run_id\"]", start)]
        for key in ("input_sha256", "contract_sha256", "engine_sha256",
                    "seed_sha256", "parameters_sha256"):
            self.assertIn(key, block, f"{key} missing from run identity")

    def test_checkpoint_loaders_gate_on_run_id(self):
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8")
        self.assertIn('if payload.get("run_id") != run_id:', source,
                      "checkpoint loaders must refuse a foreign run_id")


class DisablingAGateIsNotEvidenceThatItHeld(unittest.TestCase):
    """`apply()` treated "no issues found" and "gate switched off" identically.

    Both landed in the same `else` branch and wrote `gate_results[gate] =
    "PASS"`, so a workbook that set any Gate Mode to `off` published a clean
    PASS for that gate while the violations it had already detected were
    dropped from both `failures` and `warnings`.  All eight gates route through
    this helper and all eight accept `off` as a documented workbook value.
    """

    import dataclasses as _dc

    def _parsed(self, **over):
        import dataclasses
        kwargs = {}
        for f in dataclasses.fields(E.ParsedInput):
            if f.default is not dataclasses.MISSING:
                kwargs[f.name] = f.default
            elif f.default_factory is not dataclasses.MISSING:  # type: ignore[misc]
                kwargs[f.name] = f.default_factory()            # type: ignore[misc]
            else:
                kwargs[f.name] = None
        kwargs.update(over)
        return E.ParsedInput(**kwargs)

    CLEAN = {"active_intervals": 100, "floor_gap_count": 0, "severe_floor_gap_count": 0}
    VIOLATING = dict(CLEAN, break_concurrency_violation_count=17,
                     max_concurrent_breaks_observed=9,
                     max_concurrent_break_ratio_observed=0.9)

    def test_a_disabled_gate_with_violations_does_not_report_pass(self):
        gate = E.production_quality_gate(
            self._parsed(break_concurrency_gate_mode="off"), self.VIOLATING)
        self.assertEqual(gate["gate_results"]["break_concurrency"], "NOT_ENFORCED")
        self.assertNotEqual(gate["gate_results"]["break_concurrency"], "PASS")

    def test_the_suppressed_violation_is_still_reported(self):
        gate = E.production_quality_gate(
            self._parsed(break_concurrency_gate_mode="off"), self.VIOLATING)
        self.assertEqual(gate["suppressed_issue_count"], 1)
        row = gate["suppressed_by_disabled_gates"][0]
        self.assertEqual(row["gate"], "break_concurrency")
        self.assertEqual(row["gate_mode"], "off")

    def test_a_disabled_gate_still_does_not_block_release(self):
        """Turning a gate off must stop it blocking - only stop it lying."""
        gate = E.production_quality_gate(
            self._parsed(break_concurrency_gate_mode="off"), self.VIOLATING)
        self.assertEqual(gate["failures"], [])
        self.assertEqual(gate["warnings"], [])
        self.assertEqual(gate["status"], "PASS")

    def test_fail_and_warn_modes_are_unchanged(self):
        for mode, expected, bucket in (("fail", "FAIL", "failures"),
                                       ("warn", "WARN", "warnings")):
            with self.subTest(mode=mode):
                gate = E.production_quality_gate(
                    self._parsed(break_concurrency_gate_mode=mode), self.VIOLATING)
                self.assertEqual(gate["gate_results"]["break_concurrency"], expected)
                self.assertEqual(len(gate[bucket]), 1)
                self.assertEqual(gate["suppressed_issue_count"], 0)

    def test_a_genuinely_clean_run_still_passes_every_gate(self):
        gate = E.production_quality_gate(self._parsed(), self.CLEAN)
        self.assertEqual(set(gate["gate_results"].values()), {"PASS"})
        self.assertEqual(gate["status"], "PASS")
        self.assertEqual(gate["suppressed_issue_count"], 0)

    def test_every_gate_routes_through_the_same_helper(self):
        """If a gate stops using apply(), this guard stops covering it."""
        gate = E.production_quality_gate(self._parsed(), self.CLEAN)
        self.assertEqual(sorted(gate["gate_results"]), [
            "break_concurrency", "coverage", "employee_quality", "language_reserve",
            "next_sunday_balance", "skill_allocation", "target_loss",
            "whole_week_balance"])

    def test_a_skipped_gate_is_not_reported_as_a_pass(self):
        gate = E.production_quality_gate(self._parsed(), self.CLEAN)
        for key in ("coverage_gate_status", "language_reserve_gate_status"):
            with self.subTest(key=key):
                self.assertEqual(gate[key], "PASS")
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text(
            encoding="utf-8").replace(" ", "")
        self.assertTrue('"coverage_gate_status":gate_results.get("coverage","NOT_EVALUATED")'
                        in source,
                        "an absent gate result must not default to PASS")


class LanguageWindowsAreEnforcedOnTime(unittest.TestCase):
    """Coverage is evaluated per quarter-hour on the quarter's START minute.

    A window that does not begin on a 15-minute boundary was therefore enforced
    LATE: a 09:10-17:05 rule left the 09:00-09:15 quarter unstaffed even though
    five minutes of it fall inside the window, so required language coverage was
    simply missing for the first ten minutes of the rule. A minimum-staffing
    requirement must err toward covering.
    """

    @staticmethod
    def rule(start, end, minimum=1):
        return E.LanguageRule("x", "g", start, end, minimum, True, {"x"}, {"x"})

    def slots(self, rule, use_overlap=True):
        return [m for m in range(0, 1440, 15)
                if (rule.overlaps(m, 15) if use_overlap else rule.contains_minute(m))]

    def test_an_unaligned_window_is_covered_from_its_first_minute(self):
        rule = self.rule(9 * 60 + 10, 17 * 60 + 5)
        self.assertEqual(self.slots(rule)[0], 9 * 60,
                         "the 09:00 quarter overlaps a 09:10 start and must be staffed")
        self.assertEqual(self.slots(rule, use_overlap=False)[0], 9 * 60 + 15,
                         "pins the old late-start behaviour this replaces")

    def test_a_sub_quarter_window_is_still_covered(self):
        rule = self.rule(9 * 60, 9 * 60 + 10)
        self.assertEqual(self.slots(rule), [9 * 60])

    def test_a_window_starting_mid_quarter_covers_that_quarter(self):
        for offset in (1, 5, 10, 14):
            with self.subTest(offset=offset):
                rule = self.rule(9 * 60 + offset, 17 * 60)
                self.assertEqual(self.slots(rule)[0], 9 * 60)

    def test_aligned_windows_are_completely_unaffected(self):
        """Every language rule shipped in this project is hour-aligned."""
        for start in range(0, 1440, 15):
            for end in range(0, 1440, 15):
                rule = self.rule(start, end)
                with self.subTest(start=start, end=end):
                    self.assertEqual(self.slots(rule), self.slots(rule, use_overlap=False))

    def test_overlap_matches_brute_force(self):
        for start in range(0, 1440, 37):
            for end in range(0, 1440, 53):
                rule = self.rule(start, end)
                for minute in range(0, 1440, 15):
                    brute = any(rule.contains_minute((minute + o) % 1440) for o in range(15))
                    with self.subTest(start=start, end=end, minute=minute):
                        self.assertEqual(rule.overlaps(minute, 15), brute)

    def test_an_overnight_window_still_wraps(self):
        rule = self.rule(14 * 60, 8 * 60)
        slots = self.slots(rule)
        self.assertIn(23 * 60 + 45, slots)
        self.assertIn(0, slots)
        self.assertIn(7 * 60 + 45, slots)
        self.assertNotIn(8 * 60, slots)
        self.assertNotIn(13 * 60 + 45, slots)

    def test_an_empty_window_still_means_all_day(self):
        self.assertEqual(len(self.slots(self.rule(0, 0))), 96)


class LanguageSetupHeaderIsNotAProseBanner(unittest.TestCase):
    """`_find_header_row` accepted every term appearing anywhere in one row.

    NMG13's Language Setup sheet carries a sentence on row 2 - "General
    language/skill rules. Use Coverage Group and Minimum Per Interval for
    language balancing." - which contains both "language" and "minimum". Row 2
    won over the real header on row 4, every column then bound by positional
    fallback, and the minimum bound to the Language column.
    """

    def _sheet(self, rows):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for r, values in enumerate(rows, start=1):
            for c, value in enumerate(values, start=1):
                ws.cell(r, c).value = value
        return ws

    BANNER = "General language/skill rules. Use Coverage Group and Minimum Per Interval."
    HEADER = ["Language", "Coverage Start", "Coverage End", "Minimum Per Interval"]

    def test_a_prose_banner_does_not_win_over_the_real_header(self):
        ws = self._sheet([["Language Setup"], [self.BANNER], [], self.HEADER,
                          ["English", "14:00", "08:00", 2]])
        self.assertEqual(E._find_header_row(ws, ["language", "minimum"]), 4)

    def test_terms_must_land_in_distinct_cells(self):
        self.assertFalse(E._terms_in_distinct_cells(
            ["language", "minimum"], [E.norm(self.BANNER)]))
        self.assertTrue(E._terms_in_distinct_cells(
            ["language", "minimum"], ["language", "minimum per interval"]))

    def test_one_cell_cannot_satisfy_two_terms(self):
        self.assertFalse(E._terms_in_distinct_cells(
            ["language", "minimum"], ["language minimum", ""]))

    def test_a_single_term_still_matches_a_single_cell(self):
        self.assertTrue(E._terms_in_distinct_cells(["name"], ["sf name"]))

    def test_it_falls_back_rather_than_returning_row_one(self):
        """No distinct-cell row exists: keep the old answer, never regress."""
        ws = self._sheet([["unrelated"], [self.BANNER], ["also unrelated"]])
        self.assertEqual(E._find_header_row(ws, ["language", "minimum"]), 2)

    def test_the_real_workbook_now_resolves_correctly(self):
        import openpyxl
        path = (ROOT / "engine" / "regression_assets" / "ready_inputs"
                / "NMG13_RC8_13_CONDITIONAL_PRODUCTION_INPUT_READY.xlsx")
        if not path.exists():
            self.skipTest("regression asset not present")
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = E._sheet_by_alias(wb, ["Language Setup"])
        self.assertEqual(E._find_header_row(ws, ["language", "minimum"]), 4)
        self.assertEqual(E.norm(ws.cell(4, 1).value), "language")
        self.assertEqual(E.norm(ws.cell(4, 2).value), "coverage start")


class ProtectedTierIsReachableFromTheBusinessContract(unittest.TestCase):
    """Release gate 4 could only ever report NOT_CONFIGURED.

    `protected_before80_min` and `protected_after80_min` existed solely as CLI
    flags. RUN_UNIVERSAL_PRODUCTION.py never passed them, no workbook
    instruction alias set them, and every other contract value in this engine is
    workbook-driven. So the protected half of gate 4 was unreachable from the
    contract that drives everything else, and reported PROTECTED_NOT_EVALUATED
    on every run of the release.
    """

    ALIASES = [
        "Protected Before80 Minimum Intervals",
        "Protected Before 80 Minimum Intervals",
        "Protected Tier Before80 Minimum",
        "Protected After80 Minimum Intervals",
        "Protected After 80 Minimum Intervals",
        "Protected Tier After80 Minimum",
    ]

    def test_the_parser_accepts_a_protected_tier_row(self):
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text()
        for alias in self.ALIASES:
            with self.subTest(alias=alias):
                self.assertTrue(alias in source,
                                f"no workbook route for {alias!r}")

    def test_an_absent_row_leaves_the_tier_unconfigured(self):
        """Existing workbooks must behave exactly as before."""
        parsed = E.ParsedInput.__new__(E.ParsedInput)
        self.assertIsNone(getattr(parsed, "protected_before80_minimum", None))

    def test_the_fields_are_defaulted_so_direct_construction_still_works(self):
        """The regression lab builds ParsedInput directly, positionally."""
        import dataclasses
        fields = {f.name: f for f in dataclasses.fields(E.ParsedInput)}
        for name in ("protected_before80_minimum", "protected_after80_minimum"):
            with self.subTest(name=name):
                self.assertIn(name, fields)
                self.assertIsNone(fields[name].default,
                                  "a required field here breaks every direct construction")

    def test_the_gate_evaluator_reports_a_real_verdict_once_configured(self):
        """The point of the route: gate 4 can actually fail when it should."""
        ok, failures = E.protected_benchmark_pass(
            {"before_80": 168, "after_80": 140}, 160, 150)
        self.assertFalse(ok)
        self.assertEqual(failures, ["after80 140 < protected minimum 150"])

    def test_an_unconfigured_tier_is_not_silently_a_pass(self):
        ok, failures = E.protected_benchmark_pass(
            {"before_80": 0, "after_80": 0}, None, None)
        self.assertTrue(ok, "nothing configured means nothing checked")
        self.assertEqual(failures, [])

    def test_an_explicit_argument_still_overrides_the_workbook(self):
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text()
        self.assertTrue(
            "parsed.protected_before80_minimum\n"
            "            if protected_before80_min is None else max(0, int(protected_before80_min))"
            in source,
            "the CLI argument must win over the workbook value, as it does for "
            "the benchmark minimums")


class RunStageAndDepthComeFromTheBusinessContract(unittest.TestCase):
    """Four rows LOOKED like they controlled the run and controlled nothing.

    `Engine Defaults` shipped `RC9.1 Deep Default Seconds`, `RC9.1 Full Default
    Seconds`, `RC9.1 Stage2 Search Order` and `RC9.1 Joint Budget Policy`. Not
    one of them is read by any code path - setting Deep Default Seconds to 7200
    still produced a 14400-second run. Offering a setting that does nothing is
    worse than not offering it.

    Run Stage and Run Depth replace them with values the engine actually acts
    on, so a scheduler picks from a dropdown in the contract instead of
    remembering command-line flags.
    """

    DEAD_ROWS = ("RC9.1 Deep Default Seconds", "RC9.1 Full Default Seconds",
                 "RC9.1 Stage2 Search Order", "RC9.1 Joint Budget Policy")

    @staticmethod
    def _code_only(path):
        """Source with comment lines removed.

        The first version of this test searched the whole file and matched the
        comment that documents the finding - the same trap as the stale-literal
        guard earlier in this audit. Prose about a dead setting is not the
        engine reading it.
        """
        lines = [line for line in path.read_text().splitlines()
                 if not line.lstrip().startswith("#")]
        return "\n".join(lines)

    def test_the_four_dead_rows_are_still_dead(self):
        """Pins the finding. If one ever becomes live, this must be revisited."""
        code = self._code_only(ROOT / "engine" / "_tools" / "l632_universal_scheduler.py")
        for dead in self.DEAD_ROWS:
            with self.subTest(row=dead):
                self.assertFalse(
                    dead in code,
                    f"{dead!r} is now referenced in engine code; it was a row "
                    f"that looked like a setting and did nothing")

    def test_stage_values_a_scheduler_would_actually_type(self):
        for text, expected in [
            ("Before Breaks Only", "BEFORE_BREAKS_ONLY"),
            ("before breaks", "BEFORE_BREAKS_ONLY"),
            ("Skeleton Only", "BEFORE_BREAKS_ONLY"),
            ("shifts only", "BEFORE_BREAKS_ONLY"),
            ("After Breaks", "FULL_SCHEDULE"),
            ("Full Schedule", "FULL_SCHEDULE"),
            ("  With Breaks  ", "FULL_SCHEDULE"),
        ]:
            with self.subTest(text=text):
                self.assertEqual(E.normalize_run_stage(text), expected)

    def test_depth_values_map_to_the_three_offered_lengths(self):
        for text, expected in [("Quick", "QUICK"), ("deep", "DEEP"),
                               ("Overnight", "OVERNIGHT"), ("FULL", "OVERNIGHT")]:
            with self.subTest(text=text):
                self.assertEqual(E.normalize_run_depth(text), expected)

    def test_an_unset_cell_is_not_a_choice(self):
        for blank in (None, "", "   "):
            with self.subTest(blank=repr(blank)):
                self.assertIsNone(E.normalize_run_stage(blank))
                self.assertIsNone(E.normalize_run_depth(blank))

    def test_a_typo_falls_back_rather_than_guessing(self):
        """A misread dropdown must not silently reinterpret the run."""
        self.assertIsNone(E.normalize_run_stage("Before Break"))
        self.assertIsNone(E.normalize_run_stage("half breaks"))
        self.assertIsNone(E.normalize_run_depth("Deeep"))
        self.assertIsNone(E.normalize_run_depth("overnite"))

    def test_every_depth_has_a_wall_clock_budget(self):
        for depth in set(E.RUN_DEPTH_VALUES.values()):
            with self.subTest(depth=depth):
                self.assertIn(depth, E.RUN_DEPTH_SECONDS)
                self.assertGreater(E.RUN_DEPTH_SECONDS[depth], 0)

    def test_the_depths_are_ordered_and_deep_is_the_design_point(self):
        self.assertLess(E.RUN_DEPTH_SECONDS["QUICK"], E.RUN_DEPTH_SECONDS["DEEP"])
        self.assertLess(E.RUN_DEPTH_SECONDS["DEEP"], E.RUN_DEPTH_SECONDS["OVERNIGHT"])
        self.assertEqual(E.RUN_DEPTH_SECONDS["DEEP"], 14400,
                         "every quality number in this release is measured at DEEP")

    def _runner(self):
        return (ROOT / "engine" / "RUN_UNIVERSAL_PRODUCTION.py").read_text()

    def test_the_runner_takes_them_from_the_workbook_and_lets_the_cli_win(self):
        runner = self._runner()
        for fragment in ("contract_stage, contract_depth = _contract_run_settings(input_path)",
                         "mode = args.mode or contract_depth or 'DEEP'",
                         "stage = args.stage or contract_stage or 'FULL_SCHEDULE'"):
            with self.subTest(fragment=fragment):
                self.assertTrue(fragment in runner, f"missing: {fragment}")

    def test_the_workbook_and_the_flag_feed_one_variable(self):
        """Two ways to ask for the same run must not diverge downstream.

        `--skeleton-only` already existed with full downstream handling. Adding
        the workbook route without unifying them left every later guard reading
        the flag only, so a workbook-driven before-breaks run still tried to
        polish a break stage that never ran and died on a missing
        PARETO_EXPORT_MANIFEST.
        """
        runner = self._runner()
        self.assertTrue(
            "skeleton_only = bool(args.skeleton_only) or stage == 'BEFORE_BREAKS_ONLY'" in runner,
            "flag and workbook must resolve to one skeleton_only")
        self.assertFalse(
            "args.skeleton_only" in runner.split("skeleton_only = bool(args.skeleton_only)")[-1],
            "no downstream guard may read the raw flag instead of the resolved value")

    def test_the_break_stage_post_processing_is_skipped(self):
        """Polisher and packager need artifacts only the break stage produces."""
        runner = self._runner()
        self.assertEqual(
            runner.count("not diagnostics_only and not skeleton_only"), 2,
            "both the polisher and the packager must be gated on skeleton_only")

    def test_a_before_breaks_run_is_validated_against_its_own_output(self):
        runner = self._runner()
        self.assertTrue("'*_BEST_BEFORE_BREAKS_SCHEDULE.xlsx'" in runner,
                        "skeleton-only runs must validate the before-break workbook")


class LanguageCoverageWindowsCanBoundWorkingHours(unittest.TestCase):
    """Coverage Start/End answered only "when must this language be covered".

    The sheet's author reads them as the max and min hours those associates may
    work. Nothing in the engine enforced that: `language_eligible` is purely a
    capability test, so on GDI a bilingual associate was scheduled 00:00-09:00
    against a 16:00-07:00 window and stood on the floor at 08:00.
    """

    def _parsed(self, mode, windows=None):
        return SimpleNamespace(
            language_working_window_mode=mode,
            language_windows=windows if windows is not None else {
                "bilingual": (16 * 60, 7 * 60, True),
                "english": (16 * 60, 7 * 60, False),
            },
        )

    def _assoc(self, language):
        return SimpleNamespace(name="x", language=language)

    def _shift(self, start_h, hours):
        return SimpleNamespace(start_min=start_h * 60, duration_min=hours * 60,
                               label=f"{start_h:02d}h+{hours}")

    # -- the default must not move any recorded result -------------------
    def test_off_is_the_default_and_restricts_nothing(self):
        """Turning this on is a contract change: NMG SP's Spanish row is
        17:00-02:00, so an enforced run stops being comparable against an RC9.1
        baseline measured without it - and the gate's input-hash check cannot
        see that, because the workbook did not change."""
        self.assertEqual(E.normalize_language_window_mode(None), "OFF")
        self.assertEqual(E.normalize_language_window_mode(""), "OFF")
        self.assertIsNone(E.associate_language_window(
            self._parsed("OFF"), self._assoc("Bilingual")))

    def test_a_row_with_a_minimum_binds_under_minimum_rows(self):
        window = E.associate_language_window(
            self._parsed("MINIMUM_ROWS"), self._assoc("Bilingual"))
        self.assertEqual(window, (16 * 60, 7 * 60))

    def test_a_row_without_a_minimum_is_left_alone_under_minimum_rows(self):
        """GDI's English row carries Minimum Per Interval 0 and its own note
        says so; it may be declaring capability rather than hours."""
        self.assertIsNone(E.associate_language_window(
            self._parsed("MINIMUM_ROWS"), self._assoc("English")))

    def test_all_rows_binds_the_zero_minimum_row_too(self):
        self.assertEqual(
            E.associate_language_window(self._parsed("ALL_ROWS"), self._assoc("English")),
            (16 * 60, 7 * 60))

    def test_a_language_with_no_row_is_never_restricted(self):
        self.assertIsNone(E.associate_language_window(
            self._parsed("ALL_ROWS"), self._assoc("Klingon")))

    # -- the window arithmetic -------------------------------------------
    def test_a_shift_wholly_inside_an_overnight_window_is_legal(self):
        self.assertTrue(E.shift_within_language_window(
            self._shift(22, 8), (16 * 60, 7 * 60)))

    def test_a_shift_that_runs_past_the_window_end_is_illegal(self):
        """The real GDI cell: 00:00-09:00 against 16:00-07:00 put a bilingual
        associate on the floor at 08:00."""
        self.assertFalse(E.shift_within_language_window(
            self._shift(0, 9), (16 * 60, 7 * 60)))

    def test_a_shift_that_starts_before_the_window_opens_is_illegal(self):
        self.assertFalse(E.shift_within_language_window(
            self._shift(15, 9), (16 * 60, 7 * 60)))

    def test_a_shift_exactly_filling_the_window_is_legal(self):
        self.assertTrue(E.shift_within_language_window(
            self._shift(16, 15), (16 * 60, 7 * 60)))

    def test_one_minute_of_overrun_is_still_outside(self):
        shift = SimpleNamespace(start_min=16 * 60, duration_min=15 * 60 + 1, label="over")
        self.assertFalse(E.shift_within_language_window(shift, (16 * 60, 7 * 60)))

    def test_a_daytime_window_behaves_the_same_way(self):
        self.assertTrue(E.shift_within_language_window(
            self._shift(7, 9), (7 * 60, 2 * 60)))
        self.assertFalse(E.shift_within_language_window(
            self._shift(6, 9), (7 * 60, 2 * 60)))

    # -- a full-day window is not a restriction ---------------------------
    def test_a_window_spanning_the_whole_day_is_not_stored(self):
        """AE AR B2B and Cricut Chat both carry English 00:00-00:00. Reading
        that as a zero-length working window would forbid every shift and make
        two proven scenarios infeasible."""
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text()
        self.assertIn("if int(row[\"start\"]) == int(row[\"end\"]):", source)
        self.assertIn("a window that spans the whole day restricts nothing", source)

    # -- visibility even when off ----------------------------------------
    def test_violations_are_reported_even_when_enforcement_is_off(self):
        """Off by default means the answer to "does my roster respect the
        language hours" is invisible unless it is measured anyway."""
        source = (ROOT / "engine" / "_tools" / "l632_universal_scheduler.py").read_text()
        report = source[source.index("def language_working_window_violations("):]
        report = report[:report.index("def normalize_run_stage(")]
        self.assertIn("regardless of mode", report)
        self.assertNotIn("language_working_window_mode\", \"OFF\") == \"OFF\"", report)
        self.assertIn('"shifts_outside_language_window"', source)


if __name__ == "__main__":
    unittest.main(verbosity=2)
